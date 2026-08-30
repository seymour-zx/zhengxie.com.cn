# -*- coding: utf-8 -*-
"""
migrate_v1.py —— miniworld 数据迁移脚本（第 7 步）
========================================
读 oldworld cards.unified.xlsx → 迁移到新 schema：
1. dir_path 改名：directory/gov → topics/gov、directory/engine → topics/search、/ 保留
2. 首页卡细分 cat_name（层级 + 党派 + 工商联 + 机构，按 title 自动分类）
3. gov 人大卡 → cat_name="人大"（title 含"人民代表大会"）；其余保留
4. search 卡 cat_name 保留
5. 新增 is_ad 列（空）
输出：miniworld/assets/xlsx/cards.unified.xlsx + 迁移统计报告
"""
import openpyxl
from collections import Counter

OLD = r"D:/Universal Space/oldworld/assets/xlsx/cards.unified.xlsx"
NEW = r"D:/Universal Space/miniworld/assets/xlsx/cards.unified.xlsx"

# 党派名单
PARTIES = [
    ("中国国民党革命委员会", "民革"), ("中国民主同盟", "民盟"), ("中国民主建国会", "民建"),
    ("中国民主促进会", "民进"), ("中国农工民主党", "农工党"), ("中国致公党", "致公党"),
    ("九三学社", "九三学社"), ("台湾民主自治同盟", "台盟"),
]
# 副省级城市
SUB_PROV = ["哈尔滨","长春","沈阳","大连","济南","青岛","南京","杭州","宁波",
            "厦门","广州","深圳","武汉","成都","西安"]
# 直辖市
MUNI = ["北京","天津","上海","重庆"]

def classify_home(title):
    """首页卡分类（按 title）——修正：自治区先于区/县判断"""
    t = title or ""
    for kw, name in PARTIES:
        if kw in t:
            return name
    if "工商业联合会" in t:
        return "工商联"
    if "全国政协" in t:
        return "全国政协"
    for m in MUNI:
        if t.startswith(m) and "市政协" in t:
            return "省级政协"
    if "省政协" in t or "自治区政协" in t:
        return "省级政协"
    if "区政协" in t or "县政协" in t:
        return "县级政协"
    for c in SUB_PROV:
        if t.startswith(c) and "市政协" in t:
            return "副省级城市政协"
    if "市政协" in t or "州政协" in t or "盟政协" in t:
        return "地级政协"
    # 举报中心类 → 相关机构
    if "举报" in t or "网信办" in t:
        return "相关机构"
    return "待分类"

wb = openpyxl.load_workbook(OLD, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
header = list(rows[0])
cols = {str(h): i for i, h in enumerate(header) if h is not None}

# 新表：33 列 = 原 32 列 + is_ad
new_header = [str(h) for h in header] + ["is_ad"]

stats = Counter()
pending = []  # 待分类
gov_stats = Counter()
total = 0

wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "cards"
ws2.append(new_header)

for r in rows[1:]:
    if not r or all(c is None for c in r):
        continue
    def g(n):
        i = cols.get(n)
        return r[i] if i is not None and i < len(r) else None
    row = list(r)
    old_dir = g("dir_path") or ""
    title = g("card_title") or ""
    cat = g("cat_name") or ""
    # dir_path 改名
    if old_dir == "directory/gov":
        row[cols["dir_path"]] = "topics/gov"
        # 人大卡 → 人大类
        if "人民代表大会" in title:
            row[cols["cat_name"]] = "人大"
        gov_stats[row[cols["cat_name"]]] += 1
    elif old_dir == "directory/engine":
        row[cols["dir_path"]] = "topics/search"
    elif old_dir == "/":
        row[cols["dir_path"]] = "/"
        # 首页卡细分
        new_cat = classify_home(title)
        row[cols["cat_name"]] = new_cat
        if new_cat == "待分类":
            pending.append(title)
        stats[new_cat] += 1
    # is_ad 空
    row = list(r) + [None]
    # 用改过的 row（注意 row 的列索引与 header 一致，is_ad 在最后）
    row = list(r)
    # 重新应用修改（基于 cols 索引）
    if old_dir == "directory/gov":
        row[cols["dir_path"]] = "topics/gov"
        if "人民代表大会" in title:
            row[cols["cat_name"]] = "人大"
    elif old_dir == "directory/engine":
        row[cols["dir_path"]] = "topics/search"
    elif old_dir == "/":
        row[cols["dir_path"]] = "/"
        new_cat = classify_home(title)
        row[cols["cat_name"]] = new_cat
    row = row + [None]  # is_ad
    ws2.append(row)
    total += 1

wb2.save(NEW)
print(f"=== 迁移完成：{total} 卡 → {NEW} ===")
print(f"首页卡分类统计：")
for k, v in sorted(stats.items(), key=lambda x: -x[1]):
    print(f"  {k}: {v}")
print(f"\n待分类（需造物主定）：{len(pending)}")
for t in pending[:30]:
    print(f"  - {t}")
print(f"\ngov 分类统计：")
for k, v in sorted(gov_stats.items(), key=lambda x: -x[1]):
    print(f"  {k}: {v}")
