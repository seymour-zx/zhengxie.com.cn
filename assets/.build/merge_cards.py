# -*- coding: utf-8 -*-
"""
merge_cards.py —— 正协导航 · 各页面 cards 源 xlsx 合并器

功能：
    读取 assets/xlsx/cards-pages/ 下所有 per-page 源表
    （cards.root.xlsx / cards.gov.xlsx / cards.engine.xlsx …），
    按文件名推导各表所属频道的 dir_path，合并为单一真值源
    assets/xlsx/cards.unified.xlsx（含 dir_path 列，build_homeplus.py 直接读取）。

dir_path 推导规则（文件名驱动）：
    cards.root.xlsx       -> "/"
    cards.<channel>.xlsx  -> "directory/<channel>"   （例：cards.gov.xlsx -> directory/gov）
    root 为唯一特例；其余一律挂 directory/ 下，频道名 = 文件名去 "cards." 前缀。

用法：
    python assets/.build/merge_cards.py
    python assets/.build/merge_cards.py --src <源目录> --out <输出文件>

依赖：openpyxl
说明：
    - 仅读取每个源表的【第一个 worksheet】（worksheets[0]），忽略其余 tab
      （如部分源表附带 all_cppcc 名册 tab，避免误读导致构建报错）。
    - 列以首个源表的表头为基准（canonical）；其余源表按列名对齐到该顺序，
      缺失列补空、多余列忽略；dir_path 一律用文件名推导值覆盖，确保频道归属正确。
    - enabled 等其余列原样保留。
    - 只读源、只写 --out；不改动任何源文件。
"""
import argparse
import glob
import os
import sys

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    sys.stderr.write("缺少依赖 openpyxl，请先安装：pip install openpyxl\n")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DEFAULT_SRC = os.path.join(BASE_DIR, "assets", "xlsx", "cards-pages")
DEFAULT_OUT = os.path.join(BASE_DIR, "assets", "xlsx", "cards.unified.xlsx")

PREFIX = "cards."
SUFFIX = ".xlsx"


def channel_to_dir_path(channel):
    """cards.<channel>.xlsx -> dir_path"""
    if channel == "root":
        return "/"
    return "directory/" + channel


def main():
    ap = argparse.ArgumentParser(description="合并各页面 cards 源 xlsx 为 cards.unified.xlsx")
    ap.add_argument("--src", default=DEFAULT_SRC, help="源目录（cards.*.xlsx 所在）")
    ap.add_argument("--out", default=DEFAULT_OUT, help="输出合并文件")
    args = ap.parse_args()

    if not os.path.isdir(args.src):
        sys.stderr.write("源目录不存在: %s\n" % args.src)
        sys.exit(1)

    src_files = sorted(glob.glob(os.path.join(args.src, PREFIX + "*" + SUFFIX)))
    # 防御性排除：输出文件本不应在源目录内
    src_files = [f for f in src_files if os.path.basename(f) != "cards.unified.xlsx"]
    if not src_files:
        sys.stderr.write("源目录 %s 下未找到 %s*%s\n" % (args.src, PREFIX, SUFFIX))
        sys.exit(1)

    canonical_header = None
    all_rows = []
    per_file = []

    for f in src_files:
        base = os.path.basename(f)
        channel = base[len(PREFIX):-len(SUFFIX)]
        dir_path = channel_to_dir_path(channel)

        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.worksheets[0]  # 只取第一个 tab
        it = ws.iter_rows(values_only=True)
        header = list(next(it))
        col_idx = {name: i for i, name in enumerate(header)}

        if canonical_header is None:
            canonical_header = list(header)
            if "dir_path" not in canonical_header:
                canonical_header.insert(0, "dir_path")

        count = 0
        for r in it:
            r = list(r)
            row_dict = {name: (r[i] if i < len(r) else None) for name, i in col_idx.items()}
            row = [row_dict.get(name) for name in canonical_header]
            # 强制覆盖 dir_path，确保频道归属正确
            row[canonical_header.index("dir_path")] = dir_path
            all_rows.append(row)
            count += 1
        wb.close()
        per_file.append((base, channel, dir_path, count))
        print("  读 %-22s channel=%-8s dir_path=%-16s -> %d 行" % (base, channel, dir_path, count))

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet"
    ws_out.append(canonical_header)
    for row in all_rows:
        if len(row) < len(canonical_header):
            row = row + [None] * (len(canonical_header) - len(row))
        elif len(row) > len(canonical_header):
            row = row[:len(canonical_header)]
        ws_out.append(row)
    wb_out.save(args.out)

    print("合并完成 -> %s" % args.out)
    print("  共 %d 个源文件, %d 行, %d 列" % (len(per_file), len(all_rows), len(canonical_header)))


if __name__ == "__main__":
    main()
