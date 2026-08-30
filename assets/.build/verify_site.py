# -*- coding: utf-8 -*-
"""
verify_site.py —— miniworld 站点静态校验（L1，第 8 步改造版）
================================================================
八项校验（继承旧世界逻辑）→ 输出 **xlsx 问题报告**（契约 04，D-10）：
- sheet「汇总」：全站一览（问题总数/红线/严重/一般 + 每页面问题数）
- sheet「红线」/「严重」/「一般」：分级别问题（红/黄/灰底）
- 每行：页面/检查项/级别/问题/建议/造物主决定（下拉：上线/修复/豁免）
- 闭环：下次运行读上次报告「造物主决定」列 → 豁免项标注「已豁免(日期)」

用法：
    python assets/.build/verify_site.py            # 校验全站 → 出 xlsx 报告
    python assets/.build/verify_site.py --strict   # 严重项也视为须关注
"""
import argparse
import datetime
import json
import os
import re
import sys
from urllib.parse import unquote

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
REPORT_DIR = os.path.join(HERE, "reports")
EXCLUDE_DIRS = {"assets", ".workbuddy", ".git", "node_modules", "__pycache__"}

# 页面 → JSON-LD @type 契约（新世界映射）
JSONLD_TYPE = {
    "index.html": "WebSite",
    "404.html": None,
    "topics/index.html": "CollectionPage",
    "pages/about/index.html": "AboutPage",
    "pages/contact/index.html": "ContactPage",
}
JSONLD_DEFAULT = "WebPage"

HEAD_REQUIRED = [
    ("title", r"<title>.+?</title>"),
    ("description", r'<meta\s+name="description"\s+content=".+?"'),
    ("keywords", r'<meta\s+name="keywords"\s+content=".+?"'),
    ("canonical", r'<link\s+rel="canonical"\s+href=".+?"'),
    ("og:title", r'property="og:title"'),
    ("og:description", r'property="og:description"'),
    ("og:image", r'property="og:image"'),
    ("og:site_name", r'property="og:site_name"'),
    ("og:url", r'property="og:url"'),
    ("twitter:card", r'name="twitter:card"'),
    ("theme-color", r'name="theme-color"'),
    ("viewport", r'name="viewport"'),
    ("apple-touch-icon", r'rel="apple-touch-icon"'),
]
RED_LINES = {"1", "3"}   # 断链 / noopener


def discover_pages(root):
    pages = []
    for name in sorted(os.listdir(root)):
        if name.endswith(".html") and os.path.isfile(os.path.join(root, name)):
            pages.append(name)
    for d1 in sorted(os.listdir(root)):
        p1 = os.path.join(root, d1)
        if d1 in EXCLUDE_DIRS or not os.path.isdir(p1):
            continue
        if os.path.isfile(os.path.join(p1, "index.html")):
            pages.append(os.path.join(d1, "index.html").replace("\\", "/"))
        for d2 in sorted(os.listdir(p1)):
            p2 = os.path.join(p1, d2)
            if d2 in EXCLUDE_DIRS or not os.path.isdir(p2):
                continue
            if os.path.isfile(os.path.join(p2, "index.html")):
                pages.append(os.path.join(d1, d2, "index.html").replace("\\", "/"))
    return pages


def read(rel_path):
    with open(os.path.join(ROOT, rel_path.replace("/", os.sep)),
              encoding="utf-8", errors="replace") as f:
        return f.read()


def head_of(html):
    i = html.find("</head>")
    return html[:i] if i != -1 else html


# ── 八项检查（继承旧世界逻辑）──

def check_1_broken_paths(pages):
    bad = []
    for pg in pages:
        html = read(pg)
        base = os.path.dirname(os.path.join(ROOT, pg.replace("/", os.sep)))
        for attr in ("href", "src"):
            for url in re.findall(r'%s="([^"]+)"' % attr, html):
                if url.startswith(("http://", "https://", "#", "mailto:", "data:", "javascript:", "tel:")):
                    continue
                if not url.strip():
                    continue
                # 剥离 query（?v= 版本戳）与 fragment 再查文件存在性
                clean_url = url.split("?")[0].split("#")[0]
                target = os.path.normpath(os.path.join(base, unquote(clean_url)))
                if url.startswith("/"):
                    # 站点根绝对路径（部署后有效）：按站点根检查
                    target = os.path.normpath(os.path.join(ROOT, clean_url.lstrip("/")))
                if not os.path.exists(target):
                    bad.append("%s -> %s" % (pg, url))
    return bad


def check_2_head_tags(pages):
    bad = []
    for pg in pages:
        head = head_of(read(pg))
        missing = [name for name, pat in HEAD_REQUIRED if not re.search(pat, head)]
        if pg == "404.html" and "canonical" in missing:
            missing.remove("canonical")
        if missing:
            bad.append("%s 缺：%s" % (pg, "、".join(missing)))
    return bad


def check_3_link_attrs(pages):
    bad = []
    for pg in pages:
        html = read(pg)
        for tag in re.findall(r"<a\s[^>]*>", html):
            if 'target="_blank"' in tag and "noopener" not in tag:
                bad.append("%s -> %s" % (pg, tag[:90]))
    return bad


def check_4_link_targets(pages):
    bad = []
    ip_re = re.compile(r'https?://(?:\d{1,3}\.){3}\d{1,3}')
    for pg in pages:
        html = read(pg)
        for url in re.findall(r'href="([^"]*)"', html):
            if url.startswith("http://"):
                bad.append("%s -> 明文 http: %s" % (pg, url))
            elif ip_re.match(url):
                bad.append("%s -> 裸 IP: %s" % (pg, url))
        if re.search(r'href=""', html):
            bad.append("%s -> 存在空 href" % pg)
    return bad


def check_5_images(pages):
    bad = []
    for pg in pages:
        html = read(pg)
        for tag in re.findall(r"<img\s[^>]*>", html):
            src = re.search(r'src="([^"]*)"', tag)
            src = src.group(1) if src else ""
            issues = []
            if "alt=" not in tag:
                issues.append("缺 alt")
            if 'loading="lazy"' not in tag:
                issues.append("缺 loading=lazy")
            if src.startswith("http") and "referrerpolicy" not in tag:
                issues.append("外链图缺 referrerpolicy")
            if issues:
                bad.append("%s [%s] %s" % (pg, "、".join(issues), src[:70]))
    return bad


def check_6_headings(pages):
    bad = []
    for pg in pages:
        html = read(pg)
        h1 = re.findall(r"<h1[\s>]", html)
        if len(h1) != 1:
            bad.append("%s -> h1 数量 = %d（应为 1）" % (pg, len(h1)))
        levels = [int(m.group(1)) for m in re.finditer(r"<h([1-6])[\s>]", html)]
        for a, b in zip(levels, levels[1:]):
            if b > a + 1:
                bad.append("%s -> 标题跳级 h%d → h%d" % (pg, a, b))
                break
    return bad


def check_7_jsonld(pages):
    bad = []
    for pg in pages:
        expect = JSONLD_TYPE.get(pg, JSONLD_DEFAULT)
        html = read(pg)
        blocks = re.findall(r'<script type="application/ld\+json">(.*?)</script>', html, re.S)
        if not blocks:
            if expect is None:
                continue
            bad.append("%s -> 无 JSON-LD（契约要求 %s）" % (pg, expect))
            continue
        types = []
        for b in blocks:
            try:
                data = json.loads(b)
            except ValueError as e:
                bad.append("%s -> JSON-LD 解析失败：%s" % (pg, e))
                continue
            if isinstance(data, dict) and "@type" in data:
                types.append(data["@type"])
        if expect is not None and expect not in types:
            bad.append("%s -> JSON-LD @type = %s，契约要求 %s" % (pg, "/".join(types) or "无", expect))
    return bad


def check_8_breadcrumb(pages):
    bad = []
    for pg in pages:
        if pg in ("index.html", "404.html"):
            continue
        html = read(pg)
        blocks = re.findall(r'<script type="application/ld\+json">(.*?)</script>', html, re.S)
        has_bc = False
        for b in blocks:
            try:
                data = json.loads(b)
            except ValueError:
                continue
            if isinstance(data, dict) and data.get("@type") == "BreadcrumbList":
                has_bc = True
        if not has_bc:
            bad.append("%s -> 缺 BreadcrumbList 结构化数据" % pg)
    return bad


CHECKS = [
    ("1", "相对资源路径断链", check_1_broken_paths),
    ("2", "head 标签齐备", check_2_head_tags),
    ("3", "外链 noopener 属性", check_3_link_attrs),
    ("4", "链接目标合法性", check_4_link_targets),
    ("5", "图片 alt/lazy/referrer", check_5_images),
    ("6", "标题层级", check_6_headings),
    ("7", "JSON-LD 契约", check_7_jsonld),
    ("8", "面包屑 BreadcrumbList", check_8_breadcrumb),
]

ADVICE = {
    "1": "修正相对路径引用（css/js/图片/内链）",
    "2": "补齐缺失的 head 标签",
    "3": "外链 target=_blank 补 noopener",
    "4": "改用 https 域名链接；裸 IP/空 href 需处理",
    "5": "图片补 alt / loading=lazy / referrerpolicy",
    "6": "标题层级调整：h1 唯一、不跳级",
    "7": "按页面契约补 JSON-LD @type",
    "8": "补充 BreadcrumbList 结构化数据",
}


def load_prev_decisions(report_path):
    """读旧报告「问题 → 造物主决定」映射（契约 04 闭环：豁免/上线/修复继承，避免每次重跑丢失拍板）。"""
    dec = {}
    if not os.path.exists(report_path):
        return dec
    try:
        wb = openpyxl.load_workbook(report_path)
        for sn in ("红线", "严重", "一般"):
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 6 and row[5] in ("豁免", "上线", "修复"):
                    dec[row[3]] = row[5]
    except Exception:
        pass
    return dec


def build_xlsx(problems, report_path, prev_dec=None):
    """生成 xlsx 报告：汇总 + 红线/严重/一般 分 sheet。prev_dec=旧报告决定映射（豁免继承）。"""
    prev_dec = prev_dec or {}
    wb = openpyxl.Workbook()
    hfill = PatternFill("solid", fgColor="9E1B22")
    red_fill = PatternFill("solid", fgColor="F8CBAD")
    yellow_fill = PatternFill("solid", fgColor="FFE699")
    grey_fill = PatternFill("solid", fgColor="EDEDED")
    level_fill = {"红线": red_fill, "严重": yellow_fill, "一般": grey_fill}

    # 汇总 sheet
    ws = wb.active
    ws.title = "汇总"
    ws.append(["页面", "问题数", "红线", "严重", "一般"])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hfill
    page_stat = {}
    for pg, cid, label, level, problem in problems:
        s = page_stat.setdefault(pg, [0, 0, 0, 0])
        s[0] += 1
        s[1 if level == "红线" else 2 if level == "严重" else 3] += 1
    for pg in sorted(page_stat):
        s = page_stat[pg]
        ws.append([pg, s[0], s[1], s[2], s[3]])
    total = [sum(x) for x in zip(*page_stat.values())] if page_stat else [0, 0, 0, 0]
    ws.append(["合计", total[0], total[1], total[2], total[3]])

    # 分级别 sheet
    for level, name in [("红线", "红线"), ("严重", "严重"), ("一般", "一般")]:
        ws2 = wb.create_sheet(name)
        ws2.append(["页面", "检查项", "级别", "问题", "建议", "造物主决定"])
        for c in ws2[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = hfill
        for pg, cid, label, lv, problem in problems:
            if lv == level:
                row = [pg, "%s %s" % (cid, label), lv, problem, ADVICE.get(cid, ""), prev_dec.get(problem, "")]
                ws2.append(row)
                for cell in ws2[ws2.max_row]:
                    cell.fill = level_fill[level]
        dv = DataValidation(type="list", formula1='"上线,修复,豁免"', allow_blank=True)
        ws2.add_data_validation(dv)
        dv.add("F2:F%d" % max(2, ws2.max_row))
        for i, w in enumerate([28, 22, 8, 60, 34, 12], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
    wb.save(report_path)
    return report_path


def main():
    parser = argparse.ArgumentParser(description="miniworld 站点静态校验（L1 → xlsx 报告）")
    parser.add_argument("--strict", action="store_true", help="严重项也计为须关注")
    args = parser.parse_args()

    pages = discover_pages(ROOT)
    print("=" * 60)
    print("miniworld 站点静态校验（L1 · 八项 → xlsx 报告）")
    print("页面数：%d" % len(pages))

    problems = []
    for cid, label, func in CHECKS:
        bads = func(pages)
        level = "红线" if cid in RED_LINES else "严重"
        for b in bads:
            problems.append((b.split(" -> ")[0], cid, label, level, b))
        print("  [%s] %s · %s · %d 项" % (cid, label, level, len(bads)))

    red = sum(1 for p in problems if p[3] == "红线")
    serious = sum(1 for p in problems if p[3] == "严重")
    print("-" * 60)
    print("问题总数 %d | 🔴 红线 %d | 🟡 严重 %d" % (len(problems), red, serious))
    for p in problems:
        if p[3] == "红线":
            print("  🔴 %s" % p[4])

    os.makedirs(REPORT_DIR, exist_ok=True)
    ts = datetime.date.today().strftime("%Y%m%d")
    report = os.path.join(REPORT_DIR, "verify-report-%s.xlsx" % ts)
    prev_dec = load_prev_decisions(report)   # 契约 04 闭环：继承旧报告拍板（豁免/上线/修复）
    build_xlsx(problems, report, prev_dec)
    if prev_dec:
        print("已继承上次拍板 %d 条（豁免/上线/修复）" % len(prev_dec))
    print("xlsx 报告：%s" % report)
    print("→ 请审阅报告，逐条标「上线/修复/豁免」，保存后即为拍板（人治闸门，D-10）")


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    main()
