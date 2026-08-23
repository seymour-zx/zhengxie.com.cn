# -*- coding: utf-8 -*-
"""
collect_meta.py —— 正协导航 · SEO 元信息采集器

功能：
    递归扫描仓库内所有 index.html，提取每页的
    <title>、<meta name="keywords">、<meta name="description">，
    汇总导出为 xlsx（含「文件路径 / 标题 / 关键词 / 描述」四列）。

用法：
    python assets/.build/collect_meta.py
    python assets/.build/collect_meta.py --out 路径/report.xlsx
    python assets/.build/collect_meta.py --root 仓库根目录

依赖：openpyxl

说明：
    - 同时兼容 <meta name="x" content="..."> 与 <meta content="..." name="x"> 两种属性顺序。
    - 缺失字段记为「（无）」，不中断流程。
    - 输出 xlsx 首行冻结、列宽自适应（中文按 2 字符宽估算）。
"""

import argparse
import io
import os
import re
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("缺少依赖 openpyxl，请先安装：pip install openpyxl\n")
    sys.exit(1)


MISSING = "（无）"

# 匹配 <title>...</title>（忽略属性与大小写、跨行）
RE_TITLE = re.compile(r"<title[^>]*>(.*?)</title>", re.S | re.I)

# 匹配 <meta name="x" content="..."> 或 <meta content="..." name="x">
# 属性值允许包含除引号外的任意字符；优先 name 在前的情况，再回退 content 在前
def _build_meta_re(name):
    name_attr = r'name=["\']%s["\']' % re.escape(name)
    content_attr = r'content=["\'](.*?)["\']'
    # 顺序1: name 在前
    p1 = r"<meta[^>]+?" + name_attr + r"[^>]*?" + content_attr
    # 顺序2: content 在前
    p2 = r"<meta[^>]+?" + content_attr + r"[^>]*?" + name_attr
    return re.compile("(?:" + p1 + "|" + p2 + ")", re.S | re.I)


RE_KEYWORDS = _build_meta_re("keywords")
RE_DESCRIPTION = _build_meta_re("description")


def extract_meta(html):
    """从 HTML 文本提取 (title, keywords, description)。"""
    title_m = RE_TITLE.search(html)
    title = title_m.group(1).strip() if title_m else ""

    kw_m = RE_KEYWORDS.search(html)
    keywords = kw_m.group(1).strip() if kw_m else ""

    desc_m = RE_DESCRIPTION.search(html)
    description = desc_m.group(1).strip() if desc_m else ""

    return (
        title or MISSING,
        keywords or MISSING,
        description or MISSING,
    )


def collapse_ws(text):
    """折叠 HTML 中抽取出的多余空白（保留原语义，仅压缩连续空白与首尾）。"""
    return re.sub(r"\s+", " ", text).strip()


def collect(root):
    """返回 [(rel_path, title, keywords, description), ...] 按路径排序。"""
    results = []
    for dirpath, _dirnames, filenames in os.walk(root):
        for fn in filenames:
            if fn.lower() != "index.html":
                continue
            full = os.path.join(dirpath, fn)
            try:
                with io.open(full, encoding="utf-8", errors="replace") as fh:
                    html = fh.read()
            except OSError as e:
                sys.stderr.write("读取失败 %s: %s\n" % (full, e))
                continue
            title, keywords, description = extract_meta(html)
            rel = os.path.relpath(full, root).replace(os.sep, "/")
            results.append((rel, collapse_ws(title), collapse_ws(keywords), collapse_ws(description)))
    results.sort(key=lambda r: r[0])
    return results


def col_width(text):
    """估算列宽：中文/全角按 2，其它按 1，上限 60。"""
    width = 0
    for ch in text:
        width += 2 if ord(ch) > 0x2E7F else 1
    return min(max(width + 2, 8), 60)


def write_xlsx(results, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "SEO元信息"

    headers = ["文件路径", "标题", "关键词", "描述"]
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    header_font = Font(bold=True, color="1F4E78")
    ws.append(headers)
    for c, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in results:
        ws.append(list(row))

    # 列宽自适应（基于表头与数据）
    for ci, header in enumerate(headers, start=1):
        max_w = col_width(header)
        for r in range(2, len(results) + 2):
            val = ws.cell(row=r, column=ci).value or ""
            max_w = max(max_w, col_width(str(val)))
        ws.column_dimensions[get_column_letter(ci)].width = max_w

    # 描述列开启自动换行 + 顶端对齐，便于阅读
    for r in range(2, len(results) + 2):
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description="采集仓库内 index.html 的 title/keywords/description 并导出 xlsx")
    parser.add_argument("--root", default=os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                        help="仓库根目录（默认本脚本上两级目录，即 zhengxie.com.cn/）")
    parser.add_argument("--out", default=None, help="输出 xlsx 路径（默认 <root>/assets/.build/seo_meta_report.xlsx）")
    args = parser.parse_args()

    root = os.path.abspath(args.root)
    if not os.path.isdir(root):
        sys.stderr.write("根目录不存在：%s\n" % root)
        sys.exit(1)

    out_path = args.out or os.path.join(root, "assets", ".build", "seo_meta_report.xlsx")

    results = collect(root)
    if not results:
        sys.stderr.write("未在 %s 下找到任何 index.html\n" % root)
        sys.exit(1)

    write_xlsx(results, out_path)
    print("已采集 %d 个 index.html" % len(results))
    print("输出文件：%s" % out_path)
    for rel, title, _k, _d in results:
        print("  - %s  |  %s" % (rel, title))


if __name__ == "__main__":
    main()
