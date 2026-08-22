# -*- coding: utf-8 -*-
"""
check_links.py —— 正协导航 · 死链检测
====================================
读取 assets/xlsx/self_links.xlsx 中 links 列的 URL，逐个请求检查可达性，
输出报告 assets/py/link_report.txt。

用法：
    python assets/py/check_links.py              # 检查全部链接
    python assets/py/check_links.py --limit 5    # 只检查前 5 条（快速测试）

说明：
- links 列格式：名称,URL;名称,URL （分号分链接、逗号分名称与URL）
- 使用 HEAD 请求，遇到 405/403/501 自动降级为 GET
- 并发 8 线程，单链接超时 8 秒
"""

import argparse
import datetime
import os
import sys
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed

from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
XLSX_PATH = os.path.join(BASE_DIR, "assets", "xlsx", "self_links.xlsx")
REPORT_PATH = os.path.join(BASE_DIR, "assets", "py", "link_report.txt")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/126.0 Safari/537.36 zhengxie-linkcheck"
}


def parse_links(raw):
    """与 build.py 保持一致的解析规则：; 分链接、, 分名称与URL"""
    items = []
    for part in str(raw or "").split(";"):
        part = part.strip()
        if not part:
            continue
        if "," in part:
            name, url = part.split(",", 1)
        else:
            name, url = part, part
        url = url.strip()
        if url.startswith("http://") or url.startswith("https://"):
            items.append((name.strip() or url, url))
    return items


def collect_urls(xlsx_path, limit=None):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    urls = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or row is None:
            continue
        raw = row[7] if len(row) > 7 else None  # links 列（第 8 列）
        for name, url in parse_links(raw):
            urls.append((name, url))
        if limit is not None and len(urls) >= limit:
            break
    wb.close()
    return urls


def check(url):
    req = urllib.request.Request(url, headers=HEADERS, method="HEAD")
    try:
        with urllib.request.urlopen(req, timeout=8) as resp:
            return url, resp.status
    except urllib.error.HTTPError as e:
        # 站点不允许 HEAD 时降级为 GET
        if e.code in (405, 403, 501):
            try:
                req2 = urllib.request.Request(url, headers=HEADERS)
                with urllib.request.urlopen(req2, timeout=8) as resp:
                    return url, resp.status
            except Exception:
                return url, "ERR"
        return url, e.code
    except Exception:
        return url, "ERR"


def main():
    parser = argparse.ArgumentParser(description="正协导航死链检测")
    parser.add_argument("--limit", type=int, default=None,
                        help="只检查前 N 条 URL（快速测试）")
    args = parser.parse_args()

    urls = collect_urls(XLSX_PATH, args.limit)
    if not urls:
        sys.exit("错误：self_links.xlsx 中没有可检查的 URL。")

    print(f"开始检查 {len(urls)} 个 URL ...")
    results = []
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(check, url): (name, url) for name, url in urls}
        for fut in as_completed(futures):
            name, url = futures[fut]
            try:
                _, status = fut.result()
            except Exception:
                status = "ERR"
            results.append((status, name, url))

    ok = [r for r in results if r[0] == 200]
    bad = [r for r in results if r[0] != 200]

    lines = []
    lines.append("正协导航 · 死链检测报告")
    lines.append(f"检查时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"URL 总数: {len(results)}，正常(200): {len(ok)}，异常: {len(bad)}")
    lines.append("")
    lines.append("== 异常 / 非 200 列表 ==")
    if bad:
        for status, name, url in sorted(bad, key=lambda x: str(x[0])):
            lines.append(f"[{status}] {name}  {url}")
    else:
        lines.append("（无）")
    lines.append("")
    lines.append("== 正常列表 ==")
    for status, name, url in sorted(ok, key=lambda x: x[2]):
        lines.append(f"[{status}] {name}  {url}")

    text = "\n".join(lines)
    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        f.write(text)
    print(f"报告已写入: {REPORT_PATH}")
    print(f"正常 {len(ok)} / 异常 {len(bad)}")


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    main()
