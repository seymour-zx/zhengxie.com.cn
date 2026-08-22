# -*- coding: utf-8 -*-
"""
build.py —— 正协导航 · 站点生成器
====================================
读取 assets/xlsx/self_links.xlsx（根页独享数据源，前缀 self_ 表示独享），生成完整的 index.html。

设计原则（SEO 友好）：
- 所有导航卡片、分类按钮、链接全部内联在静态 HTML 中，不使用 JS 注入；
- 搜索引擎（百度/Google）可直接抓取全部链接与 anchor text；
- JS(main.js) 只做交互增强，禁用 JS 时页面内容完整可读可点。

用法：
    python assets/py/build.py
输出：
    index.html（站点根目录，覆盖更新）

数据表列（self_links.xlsx 第一行为表头）：
    站序 | 分类 | type | title | desc | media | tags | links
- 站序：数字，卡片按站序从小到大排列
- type：1=4行2列logo卡，2=5行横向封面卡，3=5行纵向封面卡（封面黄金比例 1.618）
- media：媒体区（逗号分隔，向后兼容旧数据）。语法：
        · `URL`                     → 仅图片（红底容器，失败移除露红底）
        · `URL,颜色`                → 图片容器内铺该背景色（给矢量/透明 logo 衬底，不改容器红底）
        · `颜色值`(#rgb/rgb()/hsl()/transparent) → 纯色块占位（无图模式，任何 CSS 合法颜色）
        · `字符,颜色`               → 文字占位 + 自定义底色（字符可为任意非 URL 非颜色文本）
        · 空/其它                   → 标题首字符 + 红渐变底（兜底）
        颜色值示例：#FFFFFF / #3A7BD5 / rgb(58,123,213) / rgba(0,0,0,.5) / hsl(210,80%,50%) / transparent
        仅按「第一个逗号」分割，颜色值内自带逗号(如 rgba(0,0,0,.5)) 不受影响。
- tags：英文逗号分隔（如 AI,免费）；分类名自动作为标签行第1个标签
- links：分号分隔链接，逗号分隔"名称与URL"（如 官网,https://x;知乎,https://z）
- 外链属性策略（target/rel）不由本表决定，而由下方 LINK_ATTR_PRESET 按**链接域名**自动匹配；
  命中预设域名（含其子域，如 www.x.com、a.b.x.com 均命中 x.com）采用对应属性，未命中一律用 DEFAULT_LINK_ATTR。

v4 变更（相对 v3）：
- 分类板块：左 logo(承担「全部」功能) + 中滑道(分类按钮) + 右「本地收藏」按钮
- 筛选板块：1行三段式（当前筛选 + 滑道 + 清除筛选）
- Hero 外搜：百度/Google/必应 主引擎按钮原位不变；下方引擎滑道放更多引擎（电商/视频/知识/开发/社交/音乐/工具/学术）
- 统一滑动行为：所有滑道（分类/筛选/引擎 + 卡片标题/描述/标签/链接行）同一套交互
  —— 只在内容真溢出时接管滚轮为左右滑（页面暂停上下滚），触屏触摸同样激活

v4.1 修订：
- media 合规 URL 校验加严：http/https + 合法主机（域名/IP/localhost），
  不合规一律视为空值 → 首字符文字占位（URL图片 / 文字占位 二选一，不叠加）
- 卡片排序：先按 type（1→2→3），再按站序；不同类型之间插入 grid-break 强制换行，
  不同类型卡片绝不同行显示
- 卡片结构重排：
  · type1 = 4行3列（媒体跨1-2行1列 / 名称1行2列 / 收藏1行3列 / 描述跨2行2-3列 / 标签3行 / 链接4行）
  · type2/3 = 5行2列（媒体跨1行 / 名称2行1列 / 收藏2行2列 / 描述3行 / 标签4行 / 链接5行）
- 卡片收藏按钮改为 SVG 星形（描边金/激活填充金），位置统一固定在名称行右端（grid 布局成员）
- 顶部 logo 保持正方形；未激活态底色改淡化红，激活态正红渐变+金环
- 「本地收藏」按钮改为与 logo 同尺寸正方形、金色系：未点击显示 ★，点击后显示「本地/收藏」两行
"""

import html
import json
import os
import re
import sys
from urllib.parse import urlparse

from openpyxl import load_workbook

# ── 路径 ──────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
# 子页（pages/*/index.html）直接以相对路径 ../../assets/ 引用根目录共享 assets，
# 不再复制 assets 进各子页目录（见 2026-08-22 调整：子页用 ../../assets/ 回退到根）。
# 因此 build 时无需 sync 子页 assets；根 assets 为唯一真源。

XLSX_PATH = os.path.join(BASE_DIR, "assets", "xlsx", "self_links.xlsx")
OUT_PATH = os.path.join(BASE_DIR, "index.html")

# ── 站点品牌（全局常量：换名只改这里一处，模板/兜底 meta 全部联动） ──
BRAND = "正协导航"
SLOGAN = "让每一次寻找，都不止于找到"
# logo 双 span 拆分：中文 4 字品牌默认前 2 字 + 后 2 字（改 BRAND 时自动联动）
_BRAND_MID = len(BRAND) // 2
BRAND_A = BRAND[:_BRAND_MID] or BRAND
BRAND_B = BRAND[_BRAND_MID:] or ""

# ── 根页元信息兜底（读不到 self_meta.json 时用此值，保证生成永不崩） ──
ROOT_META = {
    "title": f"{BRAND} - {SLOGAN}",
    "description": f"{BRAND}：全量收录的精选站点导航，覆盖常用入口、AI智能、资讯媒体、设计创意、开发技术、学习教育、效率工具、影音娱乐等分类，{SLOGAN}。",
    "keywords": f"{BRAND},网址导航,网站导航,AI工具,效率工具,政协,导航网站",
}
# directory 子页相对资源前缀（子页在 /directory/<name>/ 下，回退两级到根 assets）
DIR_ASSET_PREFIX = "../../"
DIRECTORY_ROOT = os.path.join(BASE_DIR, "directory")

# ── 搜索引擎清单（key, 显示名, 搜索URL, 是否主引擎） ──
# 主引擎（百度/Google/必应）原位不变（搜索框上方按钮）；其余进下方引擎滑道。
# 引擎清单改动只改此处一处即可（按钮与 data-url 均由此生成）。
ENGINES = [
    ("baidu",       "百度",          "https://www.baidu.com/s?wd=",                    True),
    ("google",      "Google",        "https://www.google.com/search?q=",               True),
    ("bing",        "必应",          "https://www.bing.com/search?q=",                 True),
    ("taobao",      "淘宝",          "https://s.taobao.com/search?q=",                False),
    ("jd",          "京东",          "https://search.jd.com/Search?keyword=",         False),
    ("pdd",         "拼多多",        "https://mobile.yangkeduo.com/search_result.html?search_key=", False),
    ("zhihu",       "知乎",          "https://www.zhihu.com/search?q=",               False),
    ("baike",       "百度百科",      "https://baike.baidu.com/item/",                  False),
    ("wiki",        "维基百科",      "https://zh.wikipedia.org/w/index.php?search=",   False),
    ("bilibili",    "B站",           "https://search.bilibili.com/all?keyword=",       False),
    ("douyin",      "抖音",          "https://www.douyin.com/search/",                False),
    ("github",      "GitHub",        "https://github.com/search?q=",                  False),
    ("stackoverflow", "Stack Overflow", "https://stackoverflow.com/search?q=",       False),
    ("juejin",      "掘金",          "https://juejin.cn/search?query=",                False),
    ("weibo",       "微博",          "https://s.weibo.com/weibo?q=",                   False),
    ("weixin",      "微信搜一搜",    "https://weixin.sogou.com/weixin?type=2&query=",  False),
    ("toutiao",     "头条搜索",      "https://so.toutiao.com/search?keyword=",         False),
    ("music163",    "网易云音乐",    "https://music.163.com/#/search/m/?s=",           False),
    ("youdao",      "有道翻译",      "https://dict.youdao.com/search?q=",             False),
    ("amap",        "高德地图",      "https://www.amap.com/search/?query=",           False),
    ("scholar",     "Google Scholar", "https://scholar.google.com/scholar?q=",         False),
]

# ── 站点配置（换域名只改这里一处） ──────────────────
# 末尾不要斜杠；下方所有站内绝对链接、canonical、og、JSON-LD 均由此生成。
SITE_DOMAIN = "https://zhengxie.com.cn"
# ───────────────────────────────────────────────────────────────
# 全链接属性规则（build 与子页通用，集中配置，手工增删只改这里）
# 优先级：同域 > 同族 > 营销 > 评论 > 暴露 > 默认
#   同域 (SAME_DOMAIN) ：同主域站点，原地打开（target=_self，发 Referer、传递权重）
#   同族 (SAME_FAMILY) ：品牌/姊妹站，新标签 + 仅隔离 opener（发 Referer、传递权重）
#   营销 (MARKETING)   ：广告/推广/媒体稿，新标签 + sponsored（不传递权重）
#   评论 (UGCCOMMENT)  ：论坛/社媒/评论区，新标签 + ugc（不传递权重）
#   暴露 (EXPOSED)     ：备案号/官方政务等需暴露来源，新标签 + nofollow/noopener + referrerpolicy=origin（暴露来源）
#   默认 (DEFAULT)     ：其余一切外链，新标签 + 全 nofollow/noopener/noreferrer（不传权重、不暴露来源）
# 命中逻辑：链接主机 == 域名 或 以 ".域名" 结尾（含所有子域，如 a.b.x.com 命中 x.com）。
# 增删：复制一行元组、改属性串与域名即可；要增减域名直接改对应列表。
# ───────────────────────────────────────────────────────────────
SAME_DOMAIN_ATTR = 'target="_self"'  # 同主域站点：原地打开，发 Referer、传权重
SAME_FAMILY_ATTR = 'target="_blank" rel="noopener"'
MARKETING_ATTR = 'target="_blank" rel="sponsored noopener noreferrer nofollow"'  # 当前预设空集
UGCCOMMENT_ATTR = 'target="_blank" rel="ugc noopener noreferrer nofollow"'        # 当前预设空集
EXPOSED_ATTR = 'target="_blank" rel="nofollow noopener" referrerpolicy="origin"'  # 备案号等：暴露来源
DEFAULT_LINK_ATTR = 'target="_blank" rel="nofollow noopener noreferrer"'
# 同族站点（与 SITE_DOMAIN 同主域的品牌/姊妹站）
SAME_FAMILY = ["zhengxie.info", "zhengxie.com.cn"]
# 营销站点（广告/推广/媒体稿）—— 预设空集，待后续按需要增删
MARKETING = []
# 评论站点（论坛/社媒/评论区）—— 预设空集，待后续按需要增删
UGCCOMMENT = []
# 暴露站点（备案号/官方政务等需暴露来源）
EXPOSED = ["beian.miit.gov.cn"]
EXT_LINK = EXPOSED_ATTR  # 页脚备案号等固定外链复用暴露策略

# ── 解析函数 ──────────────────────────────────────────


def parse_tags(raw):
    """tags 用英文逗号分隔，去空去重，保持顺序"""
    seen = []
    for t in str(raw or "").split(","):
        t = t.strip()
        if t and t not in seen:
            seen.append(t)
    return seen


def parse_links(raw):
    """links 用 ; 分链接、用 , 分"名称与URL"；纯URL 时名称=URL"""
    items = []
    for part in str(raw or "").split(";"):
        part = part.strip()
        if not part:
            continue
        if "," in part:
            name, url = part.split(",", 1)
        else:
            name, url = part, part
        name, url = name.strip(), url.strip()
        if url:
            items.append((name or url, url))
    return items


_DOMAIN_RE = re.compile(
    r"^(?:[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?\.)+"
    r"[a-zA-Z]{2,}$"
)
_IP_RE = re.compile(r"^\d{1,3}(?:\.\d{1,3}){3}$")


def is_url(s):
    """合规 URL：http/https 协议 + 合法主机（域名 / IP / localhost）。
    不合规（裸域名、无主机、其他协议等）一律视为空值 → 文字占位。"""
    s = (s or "").strip()
    try:
        u = urlparse(s)
    except ValueError:
        return False
    if u.scheme not in ("http", "https") or not u.netloc:
        return False
    host = u.netloc.rsplit("@", 1)[-1].split(":")[0].strip("[]")
    if host.lower() == "localhost" or _IP_RE.match(host):
        return True
    return bool(_DOMAIN_RE.match(host))


def first_char(title):
    """取标题首字符做文字logo；英文首字母大写"""
    s = (title or "").strip()
    if not s:
        return "站"
    ch = s[0]
    return ch.upper() if ch.isascii() and ch.isalpha() else ch


_COLOR_RE = re.compile(
    r"^\s*("
    r"#[0-9a-fA-F]{3,8}"                                  # #rgb / #rrggbb / #rrggbbaa
    r"|rgba?\s*\([^)]*\)"                                 # rgb() / rgba()
    r"|hsla?\s*\([^)]*\)"                                 # hsl() / hsla()
    r"|transparent"                                      # 透明
    r")\s*$"
)

# 常见 CSS 颜色关键字（小写），用于兜底识别；未列出的一律当非颜色文本
_NAMED_COLORS = {
    "transparent", "black", "white", "red", "green", "blue", "yellow", "orange",
    "purple", "gray", "grey", "cyan", "magenta", "pink", "brown", "lime",
    "navy", "teal", "silver", "gold", "maroon", "olive", "aqua", "fuchsia",
}


def _is_valid_rgblike(body):
    """校验 rgb()/rgba() 括号内参数是否合法（0~255 / 0%~100% + 可选 alpha 0~1）。"""
    parts = [p.strip() for p in body.split(",") if p.strip() != ""]
    if len(parts) not in (3, 4):
        return False
    for i, p in enumerate(parts):
        if i == 3:  # alpha
            try:
                v = float(p)
            except ValueError:
                return False
            if not (0.0 <= v <= 1.0):
                return False
        else:
            if p.endswith("%"):
                try:
                    v = float(p[:-1])
                except ValueError:
                    return False
                if not (0.0 <= v <= 100.0):
                    return False
            else:
                try:
                    v = int(p)
                except ValueError:
                    return False
                if not (0 <= v <= 255):
                    return False
    return True


def _is_valid_hsllike(body):
    """校验 hsl()/hsla() 括号内参数（H 0~360 / S,L 0%~100% + 可选 alpha）。"""
    parts = [p.strip() for p in body.split(",") if p.strip() != ""]
    if len(parts) not in (3, 4):
        return False
    # H
    try:
        h = float(parts[0])
    except ValueError:
        return False
    if not (0.0 <= h <= 360.0):
        return False
    # S, L
    for i in (1, 2):
        p = parts[i]
        if not p.endswith("%"):
            return False
        try:
            v = float(p[:-1])
        except ValueError:
            return False
        if not (0.0 <= v <= 100.0):
            return False
    # alpha
    if len(parts) == 4:
        try:
            v = float(parts[3])
        except ValueError:
            return False
        if not (0.0 <= v <= 1.0):
            return False
    return True


def is_color(s):
    """判是否为 CSS 合法颜色值，用于 media 纯色块占位。
    覆盖：#hex(3/4/6/8位) / rgb()/rgba() / hsl()/hsla() / transparent / 常见颜色名。
    对 rgb/hsl 类**校验参数范围**，非法参数（如 rgb(1,2)、rgba(...,5)）视为非法，
    避免把无效值塞进 style（浏览器会忽略，但属于「假装生效」）。"""
    s = (s or "").strip()
    if not s:
        return False
    m = _COLOR_RE.match(s)
    if not m:
        # 颜色名兜底
        return s.lower() in _NAMED_COLORS
    grp = m.group(1)
    if grp.startswith("#"):
        return True
    if grp.startswith("rgb"):
        inner = grp[grp.find("("):].strip("()")
        return _is_valid_rgblike(inner)
    if grp.startswith("hsl"):
        inner = grp[grp.find("("):].strip("()")
        return _is_valid_hsllike(inner)
    if grp == "transparent":
        return True
    return False


def build_media(media, title):
    """媒体区（media 列语法，逗号分隔，向后兼容）：
    - `URL`                  → 仅图片（红底容器，失败移除露红底）
    - `URL,颜色`             → 图片容器内铺该背景色（给矢量/透明 logo 衬底）
    - `URL,非法色/空`        → **退化为纯图**（保留 URL，不丢图）
    - `颜色值`(#hex/rgb()/hsl()/transparent/颜色名) → 纯色块占位（无图模式）
    - `合法色,任何尾巴`      → 纯色块（忽略尾巴）
    - `字符,颜色`            → 文字占位 + 自定义底色
    - `非法色 / 纯文本`      → 标题首字符 + 红渐变底（兜底）
    降级原则：任何一步颜色语句非法，都**不崩站、不丢图**，安全退到上一级可用形态。
    仅按「第一个逗号」分割，颜色值内自带逗号(如 rgba(0,0,0,.5)) 不受影响。"""
    raw = (media or "").strip()
    # 先按第一个逗号拆分（颜色值内逗号保留在 tail 中）
    if "," in raw:
        head, _, tail = raw.partition(",")
        head, tail = head.strip(), tail.strip()
    else:
        head, tail = raw, ""
    # 1) 纯 URL（head 为 URL 且无 tail）→ 纯图
    if is_url(head) and not tail:
        src = html.escape(head, quote=True)
        return (f'<div class="card__media">'
                f'<img class="card__media-img" src="{src}" alt="" '
                f'loading="lazy" decoding="async" referrerpolicy="no-referrer" '
                f'onerror="this.remove()"></div>')
    # 2) URL,合法颜色 → 图 + 底色衬底
    if is_url(head) and tail and is_color(tail):
        src = html.escape(head, quote=True)
        bg = html.escape(tail, quote=True)
        return (f'<div class="card__media" style="background:{bg}">'
                f'<img class="card__media-img" src="{src}" alt="" '
                f'loading="lazy" decoding="async" referrerpolicy="no-referrer" '
                f'onerror="this.remove()"></div>')
    # 2b) URL,非法色/空 tail → 退化为纯图（不丢图）
    if is_url(head):
        src = html.escape(head, quote=True)
        return (f'<div class="card__media">'
                f'<img class="card__media-img" src="{src}" alt="" '
                f'loading="lazy" decoding="async" referrerpolicy="no-referrer" '
                f'onerror="this.remove()"></div>')
    # 3) 纯颜色值 / 合法色带尾巴 → 纯色块占位（用 head 当色，忽略尾巴）
    if is_color(head):
        bg = html.escape(head, quote=True)
        return (f'<div class="card__media card__media--color" '
                f'style="background:{bg}"></div>')
    # 4) 字符,合法颜色 → 文字占位 + 自定义底色
    if tail and is_color(tail):
        ch = html.escape(head, quote=True)
        bg = html.escape(tail, quote=True)
        return (f'<div class="card__media" style="background:{bg}">'
                f'<span class="card__media-fallback">{ch}</span></div>')
    # 5) 兜底：标题首字符 + 红渐变底
    fb = html.escape(first_char(title))
    return (f'<div class="card__media">'
            f'<span class="card__media-fallback" aria-hidden="true">{fb}</span></div>')


def build_tags(category, tags):
    """标签行：分类名自动放第1位，其余标签随后（去重）"""
    parts = [
        f'<button type="button" class="card__tag card__tag--cat" '
        f'data-tag="{html.escape(category, quote=True)}">{html.escape(category)}</button>'
    ]
    for t in tags:
        if t == category:
            continue
        parts.append(
            f'<button type="button" class="card__tag" '
            f'data-tag="{html.escape(t, quote=True)}">{html.escape(t)}</button>'
        )
    return '<div class="card__tags">' + "".join(parts) + "</div>"


def host_of(url):
    """取 URL 主机名（去端口、转小写）；非法/无主机返回空串。"""
    try:
        host = urlparse(str(url).strip()).netloc.split(":")[0].lower()
    except (ValueError, AttributeError):
        return ""
    return host


def link_attr(url):
    """卡片/正文外链属性（优先级：同域 > 同族 > 营销 > 评论 > 暴露 > 默认）。
    同主域 → 原地打开；命中预设域名 → 对应属性；均未命中 → DEFAULT_LINK_ATTR。"""
    host = host_of(url)
    if not host:
        return DEFAULT_LINK_ATTR
    # 1) 同域（与 SITE_DOMAIN 同主域）：原地打开，发 Referer、传权重
    site_host = host_of(SITE_DOMAIN)
    if site_host and (host == site_host or host.endswith("." + site_host)):
        return SAME_DOMAIN_ATTR
    # 2) 同族
    for d in SAME_FAMILY:
        d = d.lower().strip()
        if d and (host == d or host.endswith("." + d)):
            return SAME_FAMILY_ATTR
    # 3) 营销
    for d in MARKETING:
        d = d.lower().strip()
        if d and (host == d or host.endswith("." + d)):
            return MARKETING_ATTR
    # 4) 评论
    for d in UGCCOMMENT:
        d = d.lower().strip()
        if d and (host == d or host.endswith("." + d)):
            return UGCCOMMENT_ATTR
    # 5) 暴露
    for d in EXPOSED:
        d = d.lower().strip()
        if d and (host == d or host.endswith("." + d)):
            return EXPOSED_ATTR
    # 6) 默认
    return DEFAULT_LINK_ATTR


def link_attr_footer(url):
    """页脚/导航链接属性（与卡片同源，但内链同域走 _self、备案号等暴露走 EXPOSED）。
    友情链接区(同域站)应直接用 SAME_DOMAIN_ATTR，不进此函数。"""
    return link_attr(url)


def build_links(links):
    """卡片外链：逐条按**链接域名**匹配 LINK_ATTR_PRESET 选属性串，
    命中不到的链接统一用 DEFAULT_LINK_ATTR（与 xlsx 数据无关，无需 rel 列）。"""
    parts = []
    for name, url in links:
        attr = link_attr(url)
        parts.append(
            f'<a class="card__link" href="{html.escape(url, quote=True)}" '
            f'{attr}>{html.escape(name)}<span class="card__link-arrow" aria-hidden="true">↗</span></a>'
        )
    return '<div class="card__links">' + "".join(parts) + "</div>"


FAV_SVG = (
    '<svg viewBox="0 0 24 24" aria-hidden="true" focusable="false">'
    '<path d="M12 17.27L18.18 21l-1.64-7.03L22 9.24l-7.19-.61L12 2 '
    '9.19 8.63 2 9.24l5.46 4.73L5.82 21z"/></svg>'
)


def build_card(row):
    """按 type 生成三类卡片结构之一（v4.1：收藏按钮为 SVG 星，位于名称行右端）"""
    t = str(row.get("type") or "").strip()
    cls = {"1": "card--t1", "2": "card--t2", "3": "card--t3"}.get(t, "card--t2")
    category = str(row.get("分类") or "").strip()
    title = row.get("title") or ""
    desc = row.get("desc") or ""

    media_html = build_media(row.get("media"), title)
    tags_html = build_tags(category, parse_tags(row.get("tags")))
    parsed_links = parse_links(row.get("links"))
    links_html = build_links(parsed_links)

    # 星标 key：取首个链接 URL（最稳定），无链接则退回标题
    fav_key = parsed_links[0][1] if parsed_links else (title or category or "card")
    fav_key_esc = html.escape(fav_key, quote=True)

    return (
        f'<article class="card {cls}" data-cat="{html.escape(category, quote=True)}">'
        f"{media_html}"
        f'<h3 class="card__title">{html.escape(title)}</h3>'
        f'<button type="button" class="card__fav" data-key="{fav_key_esc}" '
        f'aria-label="收藏/取消收藏" aria-pressed="false">{FAV_SVG}</button>'
        f'<p class="card__desc">{html.escape(desc)}</p>'
        f"{tags_html}{links_html}"
        f"</article>"
    )


def load_rows(xlsx_path=None):
    """读取 xlsx 全部数据行。
    v4.1 排序：先按 type（1→2→3，非法排最后），再按站序从小到大。
    xlsx_path 缺省用全局 XLSX_PATH（根页）；directory 页传入各自的 self_links.xlsx。"""
    wb = load_workbook(xlsx_path or XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row is None or all(c is None for c in row):
            continue
        if i == 0:
            header = [str(c).strip() if c is not None else "" for c in row]
            continue
        rec = {}
        for idx, col in enumerate(header):
            rec[col] = row[idx] if idx < len(row) else None
        if rec.get("站序") is None and rec.get("title") is None:
            continue
        rows.append(rec)
    wb.close()

    def type_order(r):
        t = str(r.get("type") or "").strip()
        return {"1": 1, "2": 2, "3": 3}.get(t, 9)

    def order_key(r):
        try:
            return (type_order(r), int(r.get("站序") or 0))
        except (TypeError, ValueError):
            return (type_order(r), 10**9)
    # 分类按钮顺序：原始数据从上到下首次出现（先于 type 分组排序）
    cat_order = []
    for r in rows:
        c = str(r.get("分类") or "").strip()
        if c and c not in cat_order:
            cat_order.append(c)
    rows.sort(key=order_key)
    return rows, cat_order


def build_category_buttons(cat_order):
    """分类按钮：按原始数据首现顺序去重（不随 type 分组排序）。
    v4：「全部」不再作为列表按钮——由左置 logo 承担全部功能。"""
    parts = []
    for c in cat_order:
        parts.append(
            f'<li><h2><button type="button" class="category-btn" '
            f'data-cat="{html.escape(c, quote=True)}">{html.escape(c)}</button></h2></li>'
        )
    return "".join(parts)


def build_engine_buttons():
    """生成主引擎按钮(原位) + 引擎滑道按钮(v4 新增)。
    每个按钮带 data-engine 与 data-url，JS 据此单选激活与跳转。"""
    primary_parts = []
    track_parts = []
    for i, (key, label, url, is_primary) in enumerate(ENGINES):
        active = ' class="active" aria-pressed="true"' if key == "google" else ' aria-pressed="false"'
        btn = (
            f'<button type="button" data-engine="{html.escape(key)}" '
            f'data-url="{html.escape(url, quote=True)}"{active}>'
            f'{html.escape(label)}</button>'
        )
        if is_primary:
            primary_parts.append(btn)
        else:
            track_parts.append(btn)
    return "\n        ".join(primary_parts), "\n        ".join(track_parts)


def load_meta(json_path):
    """读取 self_meta.json（页面级元信息）。文件不存在/损坏返回 {}，由调用方兜底 ROOT_META。"""
    if not json_path or not os.path.isfile(json_path):
        return {}
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except (ValueError, OSError):
        return {}


def list_directory_pages():
    """自动扫描 directory/ 下含 self_links.xlsx 的子目录。
    返回 [(name, xlsx_path, meta_path, out_path, canonical_path), ...]。
    不用手工维护列表——放一个带 self_links.xlsx 的子目录即自动生成。"""
    pages = []
    if not os.path.isdir(DIRECTORY_ROOT):
        return pages
    for name in sorted(os.listdir(DIRECTORY_ROOT)):
        d = os.path.join(DIRECTORY_ROOT, name)
        if not os.path.isdir(d):
            continue
        xlsx = os.path.join(d, "assets", "xlsx", "self_links.xlsx")
        if not os.path.isfile(xlsx):
            continue
        meta = os.path.join(d, "self_meta.json")
        out = os.path.join(d, "index.html")
        canonical = "/directory/%s/" % name
        pages.append((name, xlsx, meta, out, canonical))
    return pages


def _file_empty(path):
    """文件不存在或 0 字节 → 视为空"""
    return not os.path.isfile(path) or os.path.getsize(path) == 0


def is_empty_xlsx(path):
    """xlsx 为空：不存在 / 0字节 / 读取后无数据行 → 视为空占位"""
    if _file_empty(path):
        return True
    try:
        rows, _ = load_rows(path)
    except Exception:
        return True
    return not rows


def is_empty_meta(path):
    """meta 为空：不存在 / 0字节 / 非法JSON / 完全无字段 → 视为空占位。
    仅用于『删除空文件』判断；部分字段（填写中）不视为空，不删。"""
    return _file_empty(path) or not load_meta(path)


def build_page(category_buttons, cards_html, engine_primary, engine_track, total_cards,
               prefix="", meta=None, canonical_path="/"):
    """组装完整 index.html（静态模板，占位符替换）。

    prefix:         资源/链接路径前缀。根页=""；directory 页="../../"
    meta:           页面元信息 dict（title/description/keywords），缺失回退 ROOT_META
    canonical_path: 相对站点根的路径（"/" 或 "/directory/<name>/"），由调用方按 SITE_DOMAIN 自动拼，不来自 meta
    """
    m = dict(ROOT_META)
    if meta:
        m.update({k: v for k, v in meta.items() if v})
    return (
        PAGE_TEMPLATE.replace("{{CATEGORY_BUTTONS}}", category_buttons)
        .replace("{{CARDS}}", cards_html)
        .replace("{{ENGINE_PRIMARY}}", engine_primary)
        .replace("{{ENGINE_TRACK}}", engine_track)
        .replace("{{TOTAL_CARDS}}", str(total_cards))
        .replace("{{SITE_DOMAIN}}", SITE_DOMAIN)
        .replace("{{EXT_LINK}}", EXT_LINK)
        .replace("{{ASSET_PREFIX}}", prefix)
        .replace("{{META_TITLE}}", m["title"])
        .replace("{{META_DESC}}", m["description"])
        .replace("{{META_KEYWORDS}}", m["keywords"])
        .replace("{{CANONICAL_PATH}}", canonical_path)
        .replace("{{BRAND}}", BRAND)
        .replace("{{BRAND_A}}", BRAND_A)
        .replace("{{BRAND_B}}", BRAND_B)
        .replace("{{SLOGAN}}", SLOGAN)
    )


# ── 页面模板 ──
PAGE_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <!-- 注意：不要设全局 <meta name="referrer" content="no-referrer">——它会让百度统计/GA4 收不到来源站(referer 被禁用)。
       本站仅在卡片图片上用 referrerpolicy="no-referrer" 单独压制图片防盗链；卡片外链/引擎跳转默认发 Referer（见 link_attr 规则）。 -->
  <title>{{META_TITLE}}</title>
  <meta name="description" content="{{META_DESC}}">
  <meta name="keywords" content="{{META_KEYWORDS}}">
  <meta name="author" content="{{BRAND}}">
  <meta name="robots" content="index, follow">
  <link rel="canonical" href="{{SITE_DOMAIN}}{{CANONICAL_PATH}}">
  <!-- 社交分享 -->
  <meta property="og:title" content="{{META_TITLE}}">
  <meta property="og:description" content="{{META_DESC}}">
  <meta property="og:type" content="website">
  <meta property="og:url" content="{{SITE_DOMAIN}}{{CANONICAL_PATH}}">
  <meta property="og:image" content="{{SITE_DOMAIN}}/assets/images/logo.svg">
  <meta property="og:site_name" content="{{BRAND}}">
  <meta name="twitter:card" content="summary">
  <meta name="twitter:title" content="{{META_TITLE}}">
  <meta name="twitter:description" content="{{META_DESC}}">
  <!-- PWA / 移动端 -->
  <meta name="theme-color" content="#9E1B22" media="(prefers-color-scheme: light)">
  <meta name="theme-color" content="#0D0C0E" media="(prefers-color-scheme: dark)">
  <link rel="manifest" href="{{ASSET_PREFIX}}assets/json/manifest.json">
  <!-- 图标 -->
  <link rel="icon" type="image/svg+xml" href="{{ASSET_PREFIX}}assets/images/logo.svg">
  <link rel="apple-touch-icon" href="{{ASSET_PREFIX}}assets/images/logo.svg">
  <!-- 性能：预连接外部资源 -->
  <link rel="preconnect" href="https://www.googletagmanager.com">
  <link rel="preconnect" href="https://pagead2.googlesyndication.com">
  <link rel="preconnect" href="https://hm.baidu.com">
  <link rel="dns-prefetch" href="https://www.googletagmanager.com">
  <link rel="stylesheet" href="{{ASSET_PREFIX}}assets/css/style.css">
  <!-- 暗色模式：在 CSS 加载前同步设置，避免闪烁(FOUC)。默认明亮；仅当用户本地曾选暗色(localStorage='dark')才启用暗色 -->
  <script>
    (function(){try{var t=localStorage.getItem('zx_theme');if(t==='dark'){document.documentElement.setAttribute('data-theme','dark');}}catch(e){}})();
  </script>
  <!-- JSON-LD 结构化数据：帮助搜索引擎理解站点类型与搜索功能 -->
  <script type="application/ld+json">
  {
    "@context": "https://schema.org",
    "@type": "WebSite",
    "name": "{{META_TITLE}}",
    "alternateName": "{{META_TITLE}}",
    "url": "{{SITE_DOMAIN}}{{CANONICAL_PATH}}",
    "description": "{{META_DESC}}",
    "potentialAction": {
      "@type": "SearchAction",
      "target": {
        "@type": "EntryPoint",
        "urlTemplate": "{{SITE_DOMAIN}}/?q={search_term_string}"
      },
      "query-input": "required name=search_term_string"
    }
  }
  </script>

  <!-- ═══ 统计（GA4 + 百度统计×2，双域名各一份） ═══ -->
  <!-- Google Analytics GA4 -->
  <script async src="https://www.googletagmanager.com/gtag/js?id=G-B880S4NQVK"></script>
  <script>
    window.dataLayer = window.dataLayer || [];
    function gtag(){dataLayer.push(arguments);}
    gtag('js', new Date());
    gtag('config', 'G-B880S4NQVK');
  </script>
  <!-- 百度统计（com.cn 站 + info 站，两个站点代码合并注入） -->
  <script>
  var _hmt = _hmt || [];
  (function() {
    var ids = ["2f4df5057c929092e36a0d6357e35261", "70e38224e5ebd850150b00a19835a25f"];
    var s = document.getElementsByTagName("script")[0];
    for (var i = 0; i < ids.length; i++) {
      var hm = document.createElement("script");
      hm.src = "https://hm.baidu.com/hm.js?" + ids[i];
      s.parentNode.insertBefore(hm, s);
    }
  })();
  </script>
  <!-- ═══ Google AdSense 加载器（全页仅此一份，async 不阻塞渲染） ═══ -->
  <script async src="https://pagead2.googlesyndication.com/pagead/js/adsbygoogle.js?client=ca-pub-6434243103158481"
          crossorigin="anonymous"></script>
</head>
<body>
  <!-- 无障碍：跳到主内容 -->
  <a href="#cards-container" class="skip-link">跳到主内容</a>

  <!-- ═══ 第1行块：Hero 区 ═══ -->
  <header class="hero">
    <h1 class="hero__logo">{{BRAND}}</h1>
    <p class="hero__slogan">{{SLOGAN}}</p>
    <!-- 集合搜索引擎：主引擎按钮(原位) + 输入行 + 下方引擎滑道(扩展) -->
    <form class="hero__search" id="engine-search" action="#" role="search" aria-label="集合搜索">
      <div class="hero__engines" role="tablist" aria-label="主搜索引擎">
        {{ENGINE_PRIMARY}}
      </div>
      <div class="hero__searchrow">
        <input type="search" id="engine-input" placeholder="输入关键词，搜索全网" autocomplete="off">
        <button type="submit">搜一下</button>
      </div>
      <div class="track hero__engines-track" role="tablist" aria-label="更多引擎">
        {{ENGINE_TRACK}}
      </div>
    </form>
  </header>

  <!-- ═══ 第2行块：Google 广告位①（全宽自适应：左右零留白，撑满可用宽度给 Google 选择） ═══ -->
  <aside class="ad ad--top" aria-label="广告">
    <p class="ad__label">广告</p>
    <!-- Google AdSense 自适应展示广告（顶部专用单元，slot 与底部分开便于分位统计收益） -->
    <ins class="adsbygoogle" style="display:block"
         data-ad-client="ca-pub-6434243103158481"
         data-ad-slot="5952548493"
         data-ad-format="auto" data-full-width-responsive="true"></ins>
    <script>(adsbygoogle = window.adsbygoogle || []).push({});</script>
  </aside>

  <!-- ═══ 置顶区（整体 sticky 吸顶）：第3-5行块 ═══ -->
  <div class="sticky-top">

  <!-- ═══ 第3行块：分类导航（左 logo=全部 + 中滑道 + 右本地收藏） ═══ -->
  <nav class="category-nav" aria-label="分类筛选">
    <div class="wrap category-nav__inner">
      <button type="button" class="category-nav__logo category-btn active" data-cat="all" aria-pressed="true" aria-label="全部">
        <!-- 两份文字静态渲染，CSS 按激活态切换：激活(=全部选中)=品牌名；未激活(已选分类)显示「全部」引导返回。
             禁用 JS 时静态 HTML 带 active → 显示品牌名，SEO 无损。 -->
        <span class="logo__brand" aria-hidden="true"><span>{{BRAND_A}}</span><span>{{BRAND_B}}</span></span>
        <span class="logo__all" aria-hidden="true">全部</span>
      </button>
      <ul class="category-nav__list track" id="category-bar">
        {{CATEGORY_BUTTONS}}
      </ul>
      <button type="button" class="category-nav__fav" id="fav-toggle" aria-pressed="false" aria-label="本地收藏">
        <span class="category-nav__fav-star" aria-hidden="true">☆</span>
        <span class="category-nav__fav-text" aria-hidden="true"><span>本地</span><span>收藏</span></span>
      </button>
    </div>
  </nav>

  <!-- ═══ 第4行块：页面内搜索框（筛选站内卡片） ═══ -->
  <section class="site-search" aria-label="站内筛选">
    <div class="wrap">
      <input type="search" id="site-search-input"
             placeholder="在正协导航内筛选：标题 / 描述 / 网址关键词，回车添加筛选标签"
             autocomplete="off">
    </div>
  </section>

  <!-- ═══ 第4.5行块：结果计数（搜索框与筛选标签行之间，静态渲染总数，JS 动态更新） ═══ -->
  <section class="result-count wrap" aria-label="筛选结果统计">
    <p class="result-count__text" id="result-count">共 {{TOTAL_CARDS}} 张卡片</p>
  </section>

  <!-- ═══ 第5行块：筛选标签区（1行三段式：当前筛选 + 滑道 + 清除筛选） ═══ -->
  <section class="filter-tags wrap" id="filter-tags" aria-label="筛选标签">
    <span class="filter-tags__hint" id="filter-tags-hint">当前筛选：</span>
    <div class="track filter-tags__track" id="filter-tags-track">
      <!-- 标签 chip 由 main.js 动态插入，例：
      <span class="filter-tag"><span class="filter-tag__text">AI</span><button class="filter-tag__del" aria-label="删除该标签">×</button></span> -->
    </div>
    <button type="button" class="filter-tag__clear" id="filter-tag-clear" hidden>清除筛选</button>
  </section>

  </div><!-- /sticky-top -->

  <!-- ═══ 第6行块：导航卡片容器（build.py 生成，全部静态渲染，Grid 响应式） ═══ -->
  <main class="cards-container wrap" id="cards-container">
    {{CARDS}}
  </main>

  <!-- 空结果状态（JS 控制显隐） -->
  <div class="empty-state" id="empty-state" hidden>
    <div class="empty-state__icon" aria-hidden="true">
      <svg viewBox="0 0 64 64" width="48" height="48"><circle cx="28" cy="28" r="18" fill="none" stroke="currentColor" stroke-width="2"/><line x1="42" y1="42" x2="54" y2="54" stroke="currentColor" stroke-width="2" stroke-linecap="round"/></svg>
    </div>
    <p class="empty-state__text">没有找到匹配的结果，换个关键词或分类试试</p>
  </div>

  <!-- ═══ 第7行块：Google 广告位②（底部专用单元：slot 与顶部分开，AdSense 后台可分位统计收益） ═══ -->
  <aside class="ad ad--bottom" aria-label="广告">
    <p class="ad__label">广告</p>
    <!-- Google AdSense 自适应展示广告 -->
    <ins class="adsbygoogle" style="display:block"
         data-ad-client="ca-pub-6434243103158481"
         data-ad-slot="4856101005"
         data-ad-format="auto" data-full-width-responsive="true"></ins>
    <script>(adsbygoogle = window.adsbygoogle || []).push({});</script>
  </aside>

  <!-- ═══ 第8行块：Footer ═══ -->
  <footer class="footer">
    <div class="footer__inner wrap">
      <p class="footer__copyright">© 2026 {{BRAND}} · {{SLOGAN}}</p>
      <nav class="footer__nav" aria-label="页脚导航">
        <a href="{{ASSET_PREFIX}}">首页</a>
        <a href="{{ASSET_PREFIX}}pages/about/">关于本站</a>
        <a href="{{ASSET_PREFIX}}pages/submit/">收录申请</a>
        <a href="{{ASSET_PREFIX}}pages/contact/">联系我们</a>
        <a href="{{ASSET_PREFIX}}pages/disclaimer/">免责声明</a>
        <a href="{{ASSET_PREFIX}}pages/guide/">使用指南</a>
        <a href="{{ASSET_PREFIX}}pages/sitemap/">站点地图</a>
        <a href="{{ASSET_PREFIX}}pages/changelog/">更新日志</a>
        <a href="{{ASSET_PREFIX}}pages/privacy/">隐私政策</a>
        <a href="{{ASSET_PREFIX}}pages/overview/">网站全景</a>
        <!-- 备案号占位：当前项目托管于 GitHub Pages，无 ICP 备案，故不渲染备案链接；待迁移国内服务器完成备案后，替换粤ICP备XXXXXXXX号并取消本注释、改用以下形式：
        <a href="https://beian.miit.gov.cn/" {{EXT_LINK}}>粤ICP备XXXXXXXX号</a>
        -->
      </nav>
      <div class="footer__tools">
        <button type="button" class="footer__random" id="random-site">随机漫步</button>
        <button type="button" class="theme-toggle" id="theme-toggle" aria-label="切换深色/浅色模式" aria-pressed="false">
          <svg class="theme-toggle__sun" viewBox="0 0 24 24" aria-hidden="true"><circle cx="12" cy="12" r="5"/><line x1="12" y1="1" x2="12" y2="3"/><line x1="12" y1="21" x2="12" y2="23"/><line x1="4.2" y1="4.2" x2="5.6" y2="5.6"/><line x1="18.4" y1="18.4" x2="19.8" y2="19.8"/><line x1="1" y1="12" x2="3" y2="12"/><line x1="21" y1="12" x2="23" y2="12"/><line x1="4.2" y1="19.8" x2="5.6" y2="18.4"/><line x1="18.4" y1="5.6" x2="19.8" y2="4.2"/></svg>
          <svg class="theme-toggle__moon" viewBox="0 0 24 24" aria-hidden="true"><path d="M21 12.79A9 9 0 1 1 11.21 3 7 7 0 0 0 21 12.79z"/></svg>
        </button>
      </div>
    </div>
  </footer>

  <!-- 滚动按钮组：4 个独立按钮，按滚动位置只显示 1 个 -->
  <div class="scroll-btns" id="scroll-btns">
    <button type="button" class="scroll-btn scroll-btn--up" data-target="up" aria-label="向上滚到分类容器下方">
      <svg viewBox="0 0 24 24" aria-hidden="true">
        <path d="M12 4l-8 8h5v8h6v-8h5z" fill="currentColor"/>
      </svg>
    </button>
    <button type="button" class="scroll-btn scroll-btn--top" data-target="top" aria-label="到顶部">
      <svg viewBox="0 0 24 24" aria-hidden="true">
        <path d="M3 2.5h18v2h-18z" fill="currentColor"/>
        <path d="M12 4l-8 8h5v8h6v-8h5z" fill="currentColor"/>
      </svg>
    </button>
    <button type="button" class="scroll-btn scroll-btn--down" data-target="down" aria-label="向下滚到分类容器下方">
      <svg viewBox="0 0 24 24" aria-hidden="true">
        <path d="M12 20l-8-8h5v-8h6v8h5z" fill="currentColor"/>
      </svg>
    </button>
    <button type="button" class="scroll-btn scroll-btn--bottom" data-target="bottom" aria-label="到底部">
      <svg viewBox="0 0 24 24" aria-hidden="true">
        <path d="M12 20l-8-8h5v-8h6v8h5z" fill="currentColor"/>
        <path d="M3 19.5h18v2h-18z" fill="currentColor"/>
      </svg>
    </button>
  </div>

  <script src="{{ASSET_PREFIX}}assets/js/main.js" defer></script>
</body>
</html>
"""


def build_cards(rows):
    """生成全部卡片 HTML。v4.1：type 变化处插入 grid-break 强制换行，
    不同类型的卡片绝不在同一行显示。"""
    parts = []
    prev_type = None
    for r in rows:
        t = str(r.get("type") or "").strip()
        if prev_type is not None and t != prev_type:
            parts.append('<div class="grid-break" aria-hidden="true"></div>')
        prev_type = t
        parts.append(build_card(r))
    return "".join(parts)


def render_and_write(xlsx_path, out_path, prefix="", meta=None, canonical_path="/", label="根页"):
    """通用渲染：读取 xlsx → 生成静态页 → 写出。根页与 directory 页共用。"""
    rows, cat_order = load_rows(xlsx_path)
    if not rows:
        print(f"跳过 {label}：{xlsx_path} 无数据行。")
        return 0
    category_buttons = build_category_buttons(cat_order)
    cards = build_cards(rows)
    engine_primary, engine_track = build_engine_buttons()
    page = build_page(category_buttons, cards, engine_primary, engine_track, len(rows),
                      prefix=prefix, meta=meta, canonical_path=canonical_path)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(page)
    cats = [str(r.get("分类") or "").strip() for r in rows]
    print(f"已生成[{label}]: {out_path}  ({len(rows)} 张卡片, 分类: {' / '.join(dict.fromkeys(cats))})")
    return len(rows)


def main():
    # 1) 根页 index.html（meta 读根 self_meta.json，兜底 ROOT_META）
    #    canonical 由 SITE_DOMAIN + "/" 自动拼，不进 meta（域名与路径代码已知）
    root_meta = load_meta(os.path.join(BASE_DIR, "self_meta.json"))
    render_and_write(XLSX_PATH, OUT_PATH, prefix="", meta=root_meta,
                     canonical_path="/", label="根页")

    # 2) directory/ 子页：先清残留，再按要求生成
    if os.path.isdir(DIRECTORY_ROOT):
        # 2a) 删除所有 py 生成的 directory/<name>/index.html，避免旧页残留
        for name in sorted(os.listdir(DIRECTORY_ROOT)):
            old = os.path.join(DIRECTORY_ROOT, name, "index.html")
            if os.path.isfile(old):
                os.remove(old)
                print(f"清理: 删除旧 {old}")

        # 2b) 遍历子目录：清理空数据文件 + 生成
        #     规则：self_links.xlsx 与 self_meta.json 都齐备才生成；缺一个或都缺 → 不生成
        for name, xlsx, meta_path, out, canonical in list_directory_pages():
            # 空 self_links.xlsx（占位/未填数据）→ 删除且不生成
            if is_empty_xlsx(xlsx):
                os.remove(xlsx)
                print(f"清理: 删除空 {xlsx}")
                continue
            # self_meta.json 必填：缺失 → 不生成
            if not os.path.isfile(meta_path):
                print(f"跳过 directory/{name}：缺少 self_meta.json")
                continue
            # self_meta.json 空/非法 → 删除且不生成
            if is_empty_meta(meta_path):
                os.remove(meta_path)
                print(f"清理: 删除空 {meta_path}（空/非法）")
                continue
            # self_meta.json 字段不全（填写中）→ 不生成，但保留文件
            m = load_meta(meta_path)
            if not (m.get("title") and m.get("description") and m.get("keywords")):
                print(f"跳过 directory/{name}：self_meta.json 字段不全（保留文件）")
                continue
            render_and_write(xlsx, out, prefix=DIR_ASSET_PREFIX, meta=m,
                             canonical_path=canonical, label=f"directory/{name}")

    print(f"引擎: {len(ENGINES)} 个（主{sum(1 for e in ENGINES if e[3])} + 滑道{sum(1 for e in ENGINES if not e[3])}）")


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    main()
