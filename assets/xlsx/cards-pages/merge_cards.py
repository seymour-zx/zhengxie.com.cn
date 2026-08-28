# -*- coding: utf-8 -*-
"""
merge_cards.py —— 合并脚本【同目录】下所有 *.xlsx 为一张表

做且仅做一件事：
    读 <脚本所在目录> 下所有 *.xlsx
    合并到 <脚本所在目录>/../cards.unified.xlsx

不关心 homeplus、不推导 dir_path、不区分频道。
用法：
    python merge_cards.py
    python merge_cards.py --src <源目录> --out <输出文件>
"""
import argparse
import glob
import os
import sys

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    sys.stderr.write("缺少依赖 openpyxl，请先安装：pip install openpyxl\n")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_SRC = SCRIPT_DIR
DEFAULT_OUT = os.path.join(SCRIPT_DIR, "..", "cards.unified.xlsx")


def main():
    ap = argparse.ArgumentParser(description="合并同目录下所有 xlsx 为 cards.unified.xlsx")
    ap.add_argument("--src", default=DEFAULT_SRC, help="源目录（*.xlsx 所在，默认脚本同目录）")
    ap.add_argument("--out", default=DEFAULT_OUT, help="输出合并文件（默认上一级 cards.unified.xlsx）")
    args = ap.parse_args()

    if not os.path.isdir(args.src):
        sys.stderr.write("源目录不存在: %s\n" % args.src)
        sys.exit(1)

    src_files = sorted(glob.glob(os.path.join(args.src, "*.xlsx")))
    # 防御性排除：输出文件若已在源目录内，跳过
    src_files = [f for f in src_files if os.path.abspath(f) != os.path.abspath(args.out)]
    if not src_files:
        sys.stderr.write("源目录 %s 下未找到 *.xlsx\n" % args.src)
        sys.exit(1)

    canonical_header = None
    all_rows = []

    for f in src_files:
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.worksheets[0]  # 只取第一个 tab
        it = ws.iter_rows(values_only=True)
        header = list(next(it))
        col_idx = {name: i for i, name in enumerate(header)}

        if canonical_header is None:
            canonical_header = list(header)

        for r in it:
            r = list(r)
            row_dict = {name: (r[i] if i < len(r) else None) for name, i in col_idx.items()}
            all_rows.append([row_dict.get(name) for name in canonical_header])
        wb.close()
        print("  读 %-22s" % os.path.basename(f))

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet"
    ws_out.append(canonical_header)
    for row in all_rows:
        if len(row) < len(canonical_header):
            row = row + [None] * (len(canonical_header) - len(row))
        elif len(row) > len(canonical_header):
            row = row[:len(canonical_header)]
        ws_out.append(row)
    wb_out.save(args.out)

    print("合并完成 -> %s" % args.out)
    print("  共 %d 个源文件, %d 行, %d 列" % (len(src_files), len(all_rows), len(canonical_header)))


if __name__ == "__main__":
    main()
