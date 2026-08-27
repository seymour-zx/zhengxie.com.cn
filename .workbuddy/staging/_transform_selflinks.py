# -*- coding: utf-8 -*-
"""把 3 个在用 self_links.xlsx 的表头统一改造，生成副本到 staging（原文件不动）。
旧表头 → 统一表头（对齐候选字段清单命名），并为每个表头加批注说明。
"""
import os, openpyxl
from openpyxl.comments import Comment

SRC = [
    r'd:\Universal Space\zhengxie.com.cn\assets\xlsx\self_links.xlsx',
    r'd:\Universal Space\zhengxie.com.cn\directory\engine\assets\xlsx\self_links.xlsx',
    r'd:\Universal Space\zhengxie.com.cn\directory\gov\assets\xlsx\self_links.xlsx',
]
OUT_DIR = r'd:\Universal Space\zhengxie.com.cn\.workbuddy\staging\self_links_unified'
os.makedirs(OUT_DIR, exist_ok=True)

# 旧表头 → (统一表头, 批注说明)
MAP = {
    '站序': ('row_seq', '行序号。人工维护的顺序号，用于卡片/页面内排序（对应候选 row_id / 排序权重）。'),
    '分类': ('cat_id', '分类键。该卡片所属分类的标识（如 政协/民革/全国/直辖市）。'),
    'type': ('card_layout', '卡片版式。取值 1/2/3，表示卡片展示版式（非数据类型）。候选清单未单列，建议保留或并入 card 版式字段。'),
    'title': ('card_title', '卡片标题。可能是网站名/机构名/企业名/人名/书名/电影名甚至一条意见。'),
    'desc': ('card_desc', '卡片描述。对卡片内容的说明文字。'),
    'media': ('card_media', '卡片图标 / favicon 地址。'),
    'tags': ('card_tags', '卡片标签。多个标签用逗号分隔。'),
}
# 10 个链接槽位
for n in range(1, 11):
    MAP['link%d_name' % n] = ('link_%d_name' % n, '第 %d 个链接的显示名。' % n)
    MAP['link%d_url' % n]  = ('link_%d_url' % n,  '第 %d 个链接的地址 URL。' % n)

def tag_of(src):
    s = src.replace('\\', '/').lower()
    if '/engine/' in s:
        return 'engine'
    if '/gov/' in s:
        return 'gov'
    return 'root'

def transform(src):
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    new_headers = []
    for i, h in enumerate(headers):
        if h in MAP:
            new_h, note = MAP[h]
        else:
            new_h, note = h, '（未识别，保留原值）'
        new_headers.append(new_h)
        # 写统一表头 + 批注
        cell = ws.cell(1, i + 1)
        cell.value = new_h
        cell.comment = Comment(note, 'doc-butler')
    out = os.path.join(OUT_DIR, 'self_links.%s.unified.xlsx' % tag_of(src))
    wb.save(out)
    return out, len(headers), new_headers

if __name__ == '__main__':
    for p in SRC:
        if not os.path.exists(p):
            print('SKIP (不存在):', p); continue
        out, ncol, nh = transform(p)
        print('OK  源:', p)
        print('     副本:', out)
        print('     表头(%d):' % ncol, nh)
        print()
    print('全部完成。')
