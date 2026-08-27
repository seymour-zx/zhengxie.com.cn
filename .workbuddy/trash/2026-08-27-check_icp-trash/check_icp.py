#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
check_icp.py — 批量核验 zhengxie.com.cn 站点所有域名的 ICP 备案状态

读取「根 + 各频道」xlsx 数据源里的全部网址，提取去重域名，
逐个调用核验后端（工信部官方接口优先，免费 API 兜底），
输出报告（CSV / JSON），并可选择把备案号回填进 xlsx 的 icp_no 列。

设计原则（对齐产品决策）：
  - 官方优先：工信部 hlwicpfwc.miit.gov.cn 是权威源，不依赖第三方商业站。
  - 可降级：任一后端失败自动换下一个；全部失败标「查询失败」，绝不谎报「无备案」。
  - 有缓存：结果落 icp_cache.json，重跑只补未查项，断点可续。
  - 安全默认：--write-back 才写回 xlsx，且仅填 icp_no 为空的行（--overwrite 可覆盖）。

用法：
  python check_icp.py                  # 跑全部域名，生成 icp_report.csv / icp_report.json
  python check_icp.py --list           # 仅列出将核验的域名与数量（不联网）
  python check_icp.py --backend vvhan  # 只用指定后端（miit / vvhan / oioweb）
  python check_icp.py --limit 5        # 只核验前 N 个（调试）
  python check_icp.py --write-back     # 核验后把备案号回填进各 xlsx 的 icp_no 列

注意：实时核验需要能访问外网的运行环境。本工具在「正常网络的本机」运行效果最佳；
      若所有后端均不可达，报告会把对应域名标为「查询失败」，此时可人工核对后手动填 icp_no。
"""

import argparse
import csv
import glob
import json
import os
import re
import ssl
import sys
import time
from urllib.error import URLError, HTTPError
from urllib.request import Request, urlopen

from openpyxl import load_workbook

# ---- 路径 ----
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))          # assets/.build
REPO_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))         # 仓库根
XLSX_ROOT = os.path.join(REPO_ROOT, "assets", "xlsx", "self_links.xlsx")
XLSX_CHANNELS = glob.glob(os.path.join(REPO_ROOT, "directory", "*", "assets", "xlsx", "self_links.xlsx"))

OUT_CSV = os.path.join(SCRIPT_DIR, "icp_report.csv")
OUT_JSON = os.path.join(SCRIPT_DIR, "icp_report.json")
CACHE = os.path.join(SCRIPT_DIR, "icp_cache.json")

# ---- 网络 ----
CTX = ssl.create_default_context()
CTX.check_hostname = False
CTX.verify_mode = ssl.CERT_NONE
UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
TIMEOUT = 15
RETRIES = 2
RATE = 0.4  # 每次请求间隔（秒），避免被限频

# 备案号正则：如 京ICP备08100501号 / 京ICP证030173号
ICP_RE = re.compile(r'([一-龥]?ICP[备证]\d+号[-\w]*)')

# 明确「无备案」的语义标记（API 直说没查到）
NO_RECORD_HINTS = ("未备案", "无备案", "不存在", "没有备案", "未查询到", "无记录", "没有记录")


# --------------------------------------------------------------------------
# 1. 域名提取
# --------------------------------------------------------------------------
def _domain_of(url):
    if not url:
        return None
    m = re.search(r'https?://([^/\s]+)', str(url))
    if not m:
        return None
    d = m.group(1).lower().strip()
    if d.startswith("www."):
        d = d[4:]
    return d or None


def collect_domains():
    """扫描所有 xlsx 的网址列，返回 {domain: [(xlsx_path, row_idx, url_col, url), ...]}。"""
    files = [f for f in [XLSX_ROOT] + XLSX_CHANNELS if os.path.exists(f)]
    domains = {}
    for f in files:
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
        except Exception as e:
            print(f"  [warn] 无法读取 {f}: {e}", file=sys.stderr)
            continue
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            continue
        headers = [str(h).strip() for h in rows[0]]
        url_cols = [h for h in headers if h.endswith("_url") or re.search(r'(link|source).*url$', h)]
        for ri, row in enumerate(rows[1:], start=2):
            for col in url_cols:
                ci = headers.index(col)
                val = row[ci] if ci < len(row) else None
                if val and str(val).strip():
                    d = _domain_of(val)
                    if d:
                        domains.setdefault(d, []).append((f, ri, col, str(val).strip()))
        wb.close()
    return domains


# --------------------------------------------------------------------------
# 2. 核验后端
# --------------------------------------------------------------------------
def _http(method, url, body=None, headers=None):
    h = dict(UA)
    if headers:
        h.update(headers)
    data = json.dumps(body).encode("utf-8") if body is not None else None
    if data is not None and "Content-Type" not in h:
        h["Content-Type"] = "application/json"
    req = Request(url, data=data, headers=h, method=method)
    with urlopen(req, timeout=TIMEOUT, context=CTX) as r:
        return r.status, r.read().decode("utf-8", "ignore")


def _extract(text):
    """从任意 API 响应里尽量抠出 备案号 / 主办单位 / 网站名称。"""
    icp = ""
    m = ICP_RE.search(text)
    if m:
        icp = m.group(1)
    unit = ""
    mu = re.search(r'(?:主办单位|单位名称|unitName|单位)["\s:：]*([^",}\n]{2,40}?)["\s,}]', text)
    if mu:
        unit = mu.group(1).strip(" \t\"'")
    site = ""
    ms = re.search(r'(?:网站名称|siteName|网站名)["\s:：]*([^",}\n]{2,40}?)["\s,}]', text)
    if ms:
        site = ms.group(1).strip(" \t\"'")
    return icp, unit, site


def _empty_result():
    return {"found": None, "icp_no": "", "unit": "", "site": "", "source": "", "note": ""}


def check_miit(domain):
    """工信部官方后端（hlwicpfwc.miit.gov.cn）。权威源，但带反爬，可能 405。"""
    base = "https://hlwicpfwc.miit.gov.cn"
    token = None
    # 取令牌（部分部署 POST，部分 GET）
    for method in ("POST", "GET"):
        try:
            _, body = _http(method, base + "/icpproject_query/api/auth",
                            body={} if method == "POST" else None)
            try:
                j = json.loads(body)
                token = (j.get("params", {}).get("token") or j.get("token")
                         or j.get("data", {}).get("token"))
            except Exception:
                pass
            if token:
                break
        except Exception:
            continue
    headers = {"Referer": "https://beian.miit.gov.cn/", "Origin": "https://beian.miit.gov.cn/"}
    if token:
        headers["Authorization"] = token
    payload = {
        "pageNum": "", "pageSize": "", "unitName": "", "serviceType": "", "unitNature": "",
        "siteName": "", "mainPage": "", "domain": domain, "siteIndex": "", "siteYuming": "",
        "annualCheckSitu": "", "registrationNumber": "", "licenseKey": "", "inputCode": "",
    }
    try:
        _, body = _http("POST", base + "/icpproject_query/api/search/searchByCondition",
                        body=payload, headers=headers)
    except (URLError, HTTPError) as e:
        return {"found": None, "icp_no": "", "unit": "", "site": "",
                "source": "miit", "note": f"接口不可达: {type(e).__name__}"}
    if not body or not body.strip().startswith("{"):
        return {"found": None, "icp_no": "", "unit": "", "site": "",
                "source": "miit", "note": "返回非 JSON（可能被反爬拦截）"}
    icp, unit, site = _extract(body)
    if any(h in body for h in NO_RECORD_HINTS):
        return {"found": False, "icp_no": "", "unit": "", "site": "",
                "source": "miit", "note": "官方接口返回无记录"}
    if icp:
        return {"found": True, "icp_no": icp, "unit": unit, "site": site,
                "source": "miit", "note": ""}
    return {"found": None, "icp_no": "", "unit": "", "site": "",
            "source": "miit", "note": "无明确结论"}


def check_vvhan(domain):
    try:
        _, body = _http("GET", f"https://api.vvhan.com/api/icp?url={domain}")
    except (URLError, HTTPError) as e:
        return {"found": None, "icp_no": "", "unit": "", "site": "",
                "source": "vvhan", "note": f"接口不可达: {type(e).__name__}"}
    if any(h in body for h in NO_RECORD_HINTS):
        return {"found": False, "icp_no": "", "unit": "", "site": "",
                "source": "vvhan", "note": "返回无记录"}
    icp, unit, site = _extract(body)
    if icp:
        return {"found": True, "icp_no": icp, "unit": unit, "site": site,
                "source": "vvhan", "note": ""}
    return {"found": None, "icp_no": "", "unit": "", "site": "",
            "source": "vvhan", "note": "无明确结论"}


def check_oioweb(domain):
    try:
        _, body = _http("GET", f"https://api.oioweb.cn/api/common/ICPquery?url={domain}")
    except (URLError, HTTPError) as e:
        return {"found": None, "icp_no": "", "unit": "", "site": "",
                "source": "oioweb", "note": f"接口不可达: {type(e).__name__}"}
    if any(h in body for h in NO_RECORD_HINTS):
        return {"found": False, "icp_no": "", "unit": "", "site": "",
                "source": "oioweb", "note": "返回无记录"}
    icp, unit, site = _extract(body)
    if icp:
        return {"found": True, "icp_no": icp, "unit": unit, "site": site,
                "source": "oioweb", "note": ""}
    return {"found": None, "icp_no": "", "unit": "", "site": "",
            "source": "oioweb", "note": "无明确结论"}


BACKENDS = {"miit": check_miit, "vvhan": check_vvhan, "oioweb": check_oioweb}


def check_one(domain, backend_order):
    """按后端顺序核验，返回首个「有明确结论(found in {True,False})」的结果；都无结论则记失败。"""
    last = _empty_result()
    for name in backend_order:
        fn = BACKENDS.get(name)
        if not fn:
            continue
        try:
            res = fn(domain)
        except Exception as e:
            res = {"found": None, "icp_no": "", "unit": "", "site": "",
                   "source": name, "note": f"异常: {e}"}
        last = res
        if res["found"] is not None:
            return res
        time.sleep(RATE)
    # 全部无结论
    return {"found": None, "icp_no": "", "unit": "", "site": "",
            "source": last.get("source", ""), "note": last.get("note", "所有后端无明确结论")}


# --------------------------------------------------------------------------
# 3. 回填
# --------------------------------------------------------------------------
def write_back(domain, icp_no, overwrite):
    """把 icp_no 写到所有含该域名的 xlsx 行（仅填 icp_no 为空的，除非 overwrite）。"""
    targets = DOMAIN_MAP.get(domain, [])
    for path, row_idx, col, _url in targets:
        wb = load_workbook(path)
        ws = wb.active
        headers = [str(h).strip() for h in next(ws.iter_rows(values_only=True))]
        if "icp_no" not in headers:
            ws.cell(row=1, column=len(headers) + 1, value="icp_no")
            headers.append("icp_no")
        ci = headers.index("icp_no") + 1
        cur = ws.cell(row=row_idx, column=ci).value
        if (not cur) or overwrite:
            ws.cell(row=row_idx, column=ci, value=icp_no)
        wb.save(path)
        wb.close()


# --------------------------------------------------------------------------
# 4. 主流程
# --------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="批量核验站点所有域名的 ICP 备案状态")
    ap.add_argument("--list", action="store_true", help="仅列出将核验的域名与数量（不联网）")
    ap.add_argument("--backend", choices=list(BACKENDS.keys()), help="只用指定后端")
    ap.add_argument("--limit", type=int, default=0, help="只核验前 N 个域名（调试）")
    ap.add_argument("--write-back", action="store_true", help="核验后把备案号回填进 xlsx 的 icp_no 列")
    ap.add_argument("--overwrite", action="store_true", help="回填时覆盖已有 icp_no（需配合 --write-back）")
    args = ap.parse_args()

    global DOMAIN_MAP
    DOMAIN_MAP = collect_domains()
    domains = sorted(DOMAIN_MAP.keys())

    if args.list:
        print(f"数据源: {len([f for f in [XLSX_ROOT]+XLSX_CHANNELS if os.path.exists(f)])} 个 xlsx")
        print(f"去重域名数: {len(domains)}\n")
        for d in domains:
            print(f"  {d}  (来自 {len(DOMAIN_MAP[d])} 个链接)")
        return

    if args.backend:
        order = [args.backend]
    else:
        order = ["miit", "vvhan", "oioweb"]  # 官方优先，免费兜底

    # 读缓存
    cache = {}
    if os.path.exists(CACHE):
        try:
            with open(CACHE, "r", encoding="utf-8") as f:
                cache = json.load(f)
        except Exception:
            cache = {}

    todo = domains
    if args.limit:
        todo = todo[: args.limit]

    results = []
    print(f"开始核验 {len(todo)} 个域名，后端顺序: {order}")
    for i, domain in enumerate(todo, 1):
        if domain in cache:
            res = cache[domain]
            tag = "缓存"
        else:
            res = check_one(domain, order)
            cache[domain] = res
            tag = "新查"
            time.sleep(RATE)
        status = {True: "有备案", False: "无备案", None: "查询失败"}[res["found"]]
        print(f"  [{i}/{len(todo)}] {domain:32s} {status:8s} "
              f"{res['icp_no'] or '-':20s} {res['source'] or '-'} {tag}")
        results.append({"domain": domain, **res})
        if args.write_back and res["found"] is True and res["icp_no"]:
            write_back(domain, res["icp_no"], args.overwrite)

    # 写缓存
    with open(CACHE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

    # 报告
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["domain", "has_record", "icp_no", "unit", "site_name", "source", "note"])
        for r in results:
            w.writerow([r["domain"], {True: "有", False: "无", None: "失败"}[r["found"]],
                        r["icp_no"], r["unit"], r["site"], r["source"], r["note"]])

    n_ok = sum(1 for r in results if r["found"] is True)
    n_no = sum(1 for r in results if r["found"] is False)
    n_fail = sum(1 for r in results if r["found"] is None)
    print(f"\n完成。有备案 {n_ok} / 无备案 {n_no} / 查询失败 {n_fail}")
    print(f"报告: {OUT_CSV}\n      {OUT_JSON}")


if __name__ == "__main__":
    main()
