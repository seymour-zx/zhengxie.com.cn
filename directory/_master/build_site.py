# -*- coding: utf-8 -*-
"""
build_site.py  ——  zhengxie.com.cn 全站静态生成器（单一真值源版 · 卡片类型4）
================================================================================
读取 directory/_master/site_data.xlsx（32 列扁平行模型，每行=一个站点），
生成与线上完全一致的「生产页外壳」（搜索框 / 分类导航吸顶 / 广告位 / 页脚 /
JSON-LD / 收藏星 / 暗色切换），卡片则渲染为用户定义的「卡片类型4 · 2列×6行栅格」。

设计取舍（与用户 2026-08-26 确认）：
- 页外壳 100% 复用 assets/.build/build_homeplus.py（PAGE_TEMPLATE / build_page /
  build_jsonld / build_breadcrumb / 引擎与链接属性规则 / 收藏星 sprite），不重复造
  轮子，保证与线上页结构完全一致；本脚本只换两处：
    (1) 数据加载：site_data.xlsx 32 列 → pages / cards / links / verification / hint
    (2) 卡片渲染：卡片类型4（2列×6行栅格，见 build_card_type4 注释）
- 输出到 directory/_master/dist/，与线上零耦合；人工比对无误后复制 dist/ → 站点根。
- 不修改任何线上文件、不自动 git（治理纪律）。

卡片类型4 栅格（2 列 × 6 行）：
    列1(控件轨)       列2(内容轨)
  ┌─────────────┬──────────────────────┐
  │ (1,1) logo  │ (1,2) 卡片名称         │
  │ (2,1) 分类  │ (2~3,2) 描述(2行截断)  │
  │ (3,1) 认证  │                        │
  │ (4,1) 收藏  │ (4,2) 文字标签按钮行    │
  │ (5) 提示(法律等专家, 有列才显示) 全宽 │
  │ (6) 链接标签行(全宽)                │
  └─────────────┴──────────────────────┘
第5行「提示」对应 xlsx 可选列 card_hint：存在则渲染，不存在则该栅格位留空（不强行改 32 列模型）。
"""

import os
import shutil
import re
import html
import json
import importlib.util
from collections import OrderedDict, defaultdict
import openpyxl

# ── 路径 ──────────────────────────────────────────────
HERE = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(os.path.dirname(HERE))          # HERE=_master → 仓库根
XLSX = os.path.join(HERE, "site_data.xlsx")
DIST = os.path.join(HERE, "dist")
HP_PATH = os.path.join(BASE_DIR, "assets", ".build", "build_homeplus.py")

# ── 复用 build_homeplus.py 的生产页外壳（不重复造轮子，保证与线上页一致） ──
_spec = importlib.util.spec_from_file_location("build_homeplus_ref", HP_PATH)
hp = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(hp)
PAGE_TEMPLATE = hp.PAGE_TEMPLATE
FAV_SVG = hp.FAV_SVG
STAR_SPRITE = hp.STAR_SPRITE
ENGINES = hp.ENGINES
ROOT_META = hp.ROOT_META
BRAND = hp.BRAND
SLOGAN = hp.SLOGAN
SITE_DOMAIN = hp.SITE_DOMAIN
EXPOSED_ATTR = hp.EXPOSED_ATTR

# ── 卡片类型4 专属 CSS（生产 style.css 无此栅格，内联注入，不影响线上） ──
TYPE4_CSS = """
.card.card--t4{
  display:grid;
  grid-template-columns:60px 1fr;
  grid-template-areas:
    "logo name"
    "cat desc"
    "verify desc"
    "fav tags"
    "hint hint"
    "links links";
  gap:10px 12px;
  align-items:start;
  padding:16px;
}
.card--t4 .t4-logo{grid-area:logo;width:60px;height:60px;border-radius:10px;overflow:hidden;}
.card--t4 .t4-logo .card__media-img{max-width:100%;max-height:100%;width:auto;height:auto;}
.card--t4 .t4-name{grid-area:name;margin:0;align-self:center;font-size:17px;line-height:1.3;}
.card--t4 .t4-cat{grid-area:cat;font-size:12px;color:var(--muted,#9fb0c8);}
.card--t4 .t4-desc{grid-area:desc;margin:0;font-size:13px;color:var(--muted,#9fb0c8);
  display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden;}
.card--t4 .t4-verify{grid-area:verify;font-size:12px;line-height:1.3;}
.card--t4 .t4-verify__type{color:var(--gold,#c9a45c);font-weight:600;margin-right:4px;}
.card--t4 .t4-fav{grid-area:fav;}
.card--t4 .card__tags{grid-area:tags;display:flex;flex-wrap:wrap;gap:6px;}
.card--t4 .t4-hint{grid-area:hint;margin:0;font-size:12px;color:#b08d57;
  background:rgba(201,164,92,.10);border-left:3px solid var(--gold,#c9a45c);
  padding:6px 10px;border-radius:6px;}
.card--t4 .t4-hint__icon{margin-right:4px;font-weight:700;}
.card--t4 .card__links{grid-area:links;display:flex;flex-wrap:wrap;gap:6px;}
"""


def first(d, key, default=""):
    v = d.get(key)
    return v if v not in (None, "") else default


def parse_tags(raw):
    """card_tags 用 ; 或 , 分隔，去空去重保持顺序。"""
    seen = []
    for t in re.split(r"[;,]", str(raw or "")):
        t = t.strip()
        if t and t not in seen:
            seen.append(t)
    return seen


def load_site_data(path):
    """读取 site_data.xlsx 全部数据行（32 列扁平行）。返回 (headers, records)。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    records = []
    for r in range(2, ws.max_row + 1):
        rec = {}
        for i, h in enumerate(headers):
            v = ws.cell(row=r, column=i + 1).value
            if v not in (None, ""):
                rec[h] = v
        if rec:
            records.append(rec)
    return headers, records


def build_model(records):
    """把扁平行分组为 pages / cards / links（保留 xlsx 行序）。"""
    pages = OrderedDict()
    cards = OrderedDict()        # key=(page_id, card_id)
    card_order = defaultdict(list)
    links = defaultdict(list)
    for rec in records:
        pid = rec.get("page_id")
        if not pid:
            continue
        if pid not in pages:
            pages[pid] = {
                "page_id": pid,
                "page_title": first(rec, "page_title"),
                "page_keywords": first(rec, "page_keywords"),
                "page_description": first(rec, "page_description"),
                "slot_header_text": first(rec, "slot_header_text"),
                "slot_header_enabled": str(first(rec, "slot_header_enabled", "true")).lower() != "false",
                "slot_footer_text": first(rec, "slot_footer_text"),
                "slot_footer_enabled": str(first(rec, "slot_footer_enabled", "true")).lower() != "false",
            }
        cid = rec.get("card_id")
        if not cid:
            continue
        key = (pid, cid)
        if key not in cards:
            cards[key] = {
                "page_id": pid, "card_id": cid,
                "card_media": first(rec, "card_media"),
                "card_title": first(rec, "card_title"),
                "card_desc": first(rec, "card_desc"),
                "card_tags": first(rec, "card_tags"),
                "cat_id": first(rec, "cat_id"),
                "verification_type": first(rec, "verification_type"),
                "verification_name": first(rec, "verification_name"),
                "verification_url": first(rec, "verification_url"),
                "verification_desc": first(rec, "verification_desc"),
                "verification_enabled": str(first(rec, "verification_enabled", "false")).lower() == "true",
                "card_hint": first(rec, "card_hint"),   # 可选列：存在才渲染
            }
            card_order[pid].append(key)
        url = rec.get("url")
        if url:
            links[key].append({
                "name": first(rec, "name", url),
                "url": url,
                "desc": first(rec, "desc"),
                "media": first(rec, "media"),
                "source_type": first(rec, "source_type"),
                "verify_date": first(rec, "verify_date"),
            })
    return pages, cards, card_order, links


def card_category(c):
    """(2,1) 分类标签：优先 cat_id，否则取 card_tags 首个。"""
    cat = (c.get("cat_id") or "").strip()
    if cat:
        return cat
    tags = parse_tags(c.get("card_tags", ""))
    return tags[0] if tags else ""


def build_card_type4(c, links):
    """渲染单张「卡片类型4 · 2列×6行栅格」（字段映射见文件头注释）。"""
    cid = c["card_id"]
    title = c["card_title"]
    cat = card_category(c)

    # (1,1) logo：复用 build_homeplus.build_media（URL/纯色/首字符兜底）
    logo = hp.build_media(c.get("card_media", ""), title)
    logo = logo.replace('class="card__media"', 'class="card__media t4-logo"', 1)

    # (1,2) 名称
    name = f'<h3 class="card__title t4-name">{html.escape(title)}</h3>'

    # (2,1) 分类标签
    cat_html = f'<span class="t4-cat">{html.escape(cat)}</span>' if cat else ""

    # (2~3,2) 描述（2 行占位，超出截断，由 CSS -webkit-line-clamp 控制）
    desc = f'<p class="card__desc t4-desc">{html.escape(c["card_desc"])}</p>' if c["card_desc"] else ""

    # (3,1) verification 链接标签（仅启用且有 url 时渲染；备案号走公开来源策略）
    verify = ""
    if c["verification_enabled"] and c["verification_url"]:
        vtype = html.escape(c["verification_type"] or "认证")
        vdesc = html.escape(c["verification_desc"] or c["verification_name"] or "")
        verify = (f'<a class="t4-verify" href="{html.escape(c["verification_url"], quote=True)}" '
                  f'{EXPOSED_ATTR} title="{vdesc}">'
                  f'<span class="t4-verify__type">{vtype}</span>{vdesc}</a>')

    # (4,1) 收藏按钮（SVG 星，localStorage 由 main.js 接管；key 取首个链接 url）
    key = links[0]["url"] if links else (title or cid)
    fav = (f'<button type="button" class="card__fav t4-fav" data-key="{html.escape(key, quote=True)}" '
           f'aria-label="收藏/取消收藏" aria-pressed="false">{FAV_SVG}</button>')

    # (4,2) 文字标签按钮行（card_tags 拆分；data-tag 供筛选）
    tag_html = ""
    tags = parse_tags(c["card_tags"])
    if tags:
        parts = [
            f'<button type="button" class="card__tag" data-tag="{html.escape(t, quote=True)}">{html.escape(t)}</button>'
            for t in tags
        ]
        tag_html = '<div class="card__tags">' + "".join(parts) + "</div>"

    # (5) 提示（法律等专家说的提示文本）—— 仅当 xlsx 存在 card_hint 列且有值
    hint = ""
    if c.get("card_hint"):
        hint = (f'<p class="t4-hint"><span class="t4-hint__icon" aria-hidden="true">ⓘ</span>'
                f'{html.escape(c["card_hint"])}</p>')

    # (6) 链接标签行（name → url；属性按域名自动匹配，由 build_homeplus.build_links 处理）
    links_html = ""
    if links:
        parsed = [(l["name"] or l["url"], l["url"]) for l in links]
        links_html = hp.build_links(parsed)

    return (f'<article class="card card--t4" data-cat="{html.escape(cat, quote=True)}">'
            f'{logo}{name}{cat_html}{desc}{verify}{fav}{tag_html}{hint}{links_html}'
            f'</article>')


def build_jsonld_rows(card_links):
    """把卡片链接拼成 build_homeplus.build_jsonld 能消费的伪行（取首个外链进 ItemList）。"""
    rows = []
    for ls in card_links:
        if ls:
            rows.append({"links": ";".join(f'{l["name"]},{l["url"]}' for l in ls)})
    return rows


def render_page(pid, page, cards, cards_for_page, links):
    """组装单个页面：复用 build_homeplus.build_page（生产页外壳）。"""
    is_root = pid in ("home", "")
    # dist/ 作为站点根镜像：首页在 dist/，频道页在 dist/<pid>/（仅深一层），
    # 共享资源统一放 dist/assets/ → 首页前缀 ""、频道页前缀 "../"
    prefix = "" if is_root else "../"
    canonical = "/" if is_root else f"/{pid}/"

    cat_order = []
    card_htmls = []
    card_links = []
    for key in cards_for_page:
        c = cards[key]
        ls = links.get(key, [])
        card_htmls.append(build_card_type4(c, ls))
        card_links.append(ls)
        cc = card_category(c)
        if cc and cc not in cat_order:
            cat_order.append(cc)

    category_buttons = hp.build_category_buttons(cat_order)
    engine_primary, engine_track = hp.build_engine_buttons()

    meta = {k: page.get(k, "") for k in ("page_title", "page_keywords", "page_description")}
    meta = {**ROOT_META, **{k: v for k, v in meta.items() if v}}

    # hero 下半部：根页=集合搜索框；频道页=专题介绍块；均可在上方追加 slot_header 副标题
    sub = page["slot_header_text"] if page["slot_header_enabled"] else ""
    sub_html = f'<p class="hero__sub">{html.escape(sub)}</p>\n' if sub else ""
    if is_root:
        hero_search = sub_html + hp.HERO_SEARCH_BLOCK
    else:
        hero_search = sub_html + hp.build_channel_intro(meta)

    channel_name = meta["title"].replace(" - 正协导航", "").replace(" -正协导航", "").strip()
    page_html = hp.build_page(
        category_buttons, "".join(card_htmls), engine_primary, engine_track,
        len(card_htmls), prefix=prefix, meta=meta, canonical_path=canonical,
        hero_search=hero_search, rows=build_jsonld_rows(card_links),
        channel_name=channel_name,
    )

    # 注入卡片类型4 专属 CSS（共享 style.css 之后，保证覆盖）
    page_html = page_html.replace("</head>", f"<style>{TYPE4_CSS}</style></head>")
    # 页脚 slot（slot_footer_text，可选）
    if page["slot_footer_enabled"] and page["slot_footer_text"]:
        page_html = page_html.replace(
            "</footer>",
            f'<p class="footer__note">{html.escape(page["slot_footer_text"])}</p></footer>',
        )
    return page_html, canonical


def main():
    os.makedirs(DIST, exist_ok=True)

    # 把共享资源与手写页复制进 dist/，使 dist/ 成为自包含站点根镜像
    # （首页 assets/、频道页 ../assets/ 均精确指向 dist/assets/；页脚 ../pages/ 指向 dist/pages/）
    for src_name, dst_name in (("assets", "assets"), ("pages", "pages")):
        src = os.path.join(BASE_DIR, src_name)
        dst = os.path.join(DIST, dst_name)
        if os.path.isdir(src):
            shutil.copytree(src, dst, dirs_exist_ok=True)
            print(f"同步[{src_name}] → {dst}")

    headers, records = load_site_data(XLSX)
    pages, cards, card_order, links = build_model(records)

    written = []
    for pid in pages:
        page = pages[pid]
        # 跳过 0 卡片页面：多为手写 pages/ 子页（about/submit/...）或空定义，不覆盖、不生成
        if not card_order[pid]:
            print(f"跳过[{pid}]: 无卡片（手写页或空定义），不生成。")
            continue
        page_html, canonical = render_page(pid, page, cards, card_order[pid], links)
        out_rel = "index.html" if pid in ("home", "") else os.path.join(pid, "index.html")
        out_abs = os.path.join(DIST, out_rel)
        os.makedirs(os.path.dirname(out_abs), exist_ok=True)
        with open(out_abs, "w", encoding="utf-8") as f:
            f.write(page_html)
        n = len(card_order[pid])
        written.append((pid, canonical, out_rel, n))
        print(f"生成[{pid}]: {out_abs}  ({n} 张卡片)")

    # sitemap.xml
    locs = [
        f'  <url><loc>{html.escape(SITE_DOMAIN + canonical)}</loc></url>'
        for _pid, canonical, _rel, _n in written
    ]
    sm = ('<?xml version="1.0" encoding="UTF-8"?>\n'
          '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
          + "\n".join(locs) + "\n</urlset>\n")
    with open(os.path.join(DIST, "sitemap.xml"), "w", encoding="utf-8") as f:
        f.write(sm)

    print(f"\n源表列数: {len(headers)} | 数据行: {len(records)}")
    print(f"页面: {len(pages)} | 卡片: {sum(len(v) for v in card_order.values())} | 主链接: {sum(len(v) for v in links.values())}")
    print(f"已生成 {len(written)} 个页面 + sitemap.xml 至: {DIST}")
    for pid, canonical, rel, n in written:
        print(f"  + {rel}  ({canonical})")


if __name__ == "__main__":
    try:
        import sys
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    main()
