import openpyxl

# ONE-SHOT 转换：固定从 site_data.bak.xlsx（原始 36 列）读取，写出 32 列 site_data.xlsx。
# 幂等安全：重复运行都从原始备份重建，不会因"跑两次"丢失 attest/slot。
SRC = r'd:\Universal Space\zhengxie.com.cn\directory\_master\site_data.bak.xlsx'
OUT = r'd:\Universal Space\zhengxie.com.cn\directory\_master\site_data.xlsx'

NEW_HEADERS = [
    'row_id','page_id','page_title','page_keywords','page_description',
    'slot_header_text','slot_header_enabled','slot_footer_text','slot_footer_enabled',
    'cat_id','card_id','card_media','card_title','card_desc','card_tags',
    'verification_type','verification_name','verification_url','verification_desc','verification_enabled',
    'link_id','name','url','desc','media',
    'source_type','verify_date','verify_channel','link_status','review_cycle',
    'created_at','updated_at',
]

EMPTY = {h:'' for h in NEW_HEADERS}

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb.active
H = [c.value for c in ws[1]]
rows=[]
for r in range(2, ws.max_row+1):
    rows.append({H[i]: ws.cell(r,i+1).value for i in range(len(H))})

# ---- 1. collect attest data per card_id (flatten: keep FIRST only) ----
attest = {}
multi_warn = []
for rec in rows:
    cid = rec.get('attest_card_id')
    if not cid:
        continue
    tags = rec.get('tags') or ''
    vtype = tags.strip() if tags else ''
    entry = {
        'verification_type': vtype,
        'verification_name': '',
        'verification_url': rec.get('url') or '',
        'verification_desc': (rec.get('desc') or '').strip(),
        'verification_enabled': 'true',
    }
    if cid in attest:
        multi_warn.append(cid)
    else:
        attest[cid] = entry

# ---- 2. merge page-level fields per page_id (from url-empty, non-attest rows) ----
page_data = {}
for rec in rows:
    pid = rec.get('page_id')
    if not pid or rec.get('url') or rec.get('attest_card_id'):
        continue
    d = page_data.setdefault(pid, {})
    for k in ['page_title','page_keywords','page_description']:
        if rec.get(k):
            d.setdefault(k, rec[k])
    if rec.get('slot_key'):
        d['slot_header_text'] = rec.get('slot_text')
        d['slot_header_enabled'] = rec.get('slot_enabled')

# ---- 3. emit rows (page def captured from page_data on first occurrence) ----
out = []
emitted_pages = set()
seen_card = set()

def make_page_row(pid, d):
    r = dict(EMPTY)
    r['page_id'] = pid
    r['page_title'] = d.get('page_title','')
    r['page_keywords'] = d.get('page_keywords','')
    r['page_description'] = d.get('page_description','')
    r['slot_header_text'] = d.get('slot_header_text','')
    r['slot_header_enabled'] = d.get('slot_header_enabled','')
    return r

def make_card_row(rec):
    r = dict(EMPTY)
    cid = rec.get('card_id')
    r['page_id'] = rec.get('page_id') or ''
    if cid and cid not in seen_card:
        seen_card.add(cid)
        r['card_id'] = cid
        r['card_media'] = rec.get('card_media','')
        r['card_title'] = rec.get('card_title','')
        r['card_desc'] = rec.get('card_desc','')
        r['card_tags'] = rec.get('card_tags','')
        r['cat_id'] = rec.get('cat_id','')
        if cid in attest:
            r['verification_type'] = attest[cid]['verification_type']
            r['verification_name'] = attest[cid]['verification_name']
            r['verification_url'] = attest[cid]['verification_url']
            r['verification_desc'] = attest[cid]['verification_desc']
            r['verification_enabled'] = attest[cid]['verification_enabled']
    else:
        r['card_id'] = cid or ''
    r['link_id'] = rec.get('link_id','')
    r['name'] = rec.get('name','')
    r['url'] = rec.get('url','')
    r['desc'] = rec.get('desc','')
    r['media'] = rec.get('media','')
    r['source_type'] = rec.get('source_type','')
    r['verify_date'] = rec.get('verify_date','')
    r['verify_channel'] = rec.get('verify_channel','')
    r['link_status'] = rec.get('link_status','')
    r['review_cycle'] = rec.get('review_cycle','')
    r['created_at'] = rec.get('created_at','')
    r['updated_at'] = rec.get('updated_at','')
    return r

for rec in rows:
    if rec.get('attest_card_id'):
        continue
    pid = rec.get('page_id')
    if pid and pid not in emitted_pages:
        emitted_pages.add(pid)
        out.append(make_page_row(pid, page_data.get(pid, {})))
        if rec.get('url'):
            out.append(make_card_row(rec))
    elif rec.get('url'):
        out.append(make_card_row(rec))
    # url-empty page def row whose page already emitted -> skip

for i,row in enumerate(out,1):
    row['row_id']=i

wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = '全量数据'
ws2.append(NEW_HEADERS)
for row in out:
    ws2.append([row[h] for h in NEW_HEADERS])
wb2.save(OUT)

print('WROTE rows (excl header):', len(out))
print('NEW columns:', len(NEW_HEADERS))
print('Pages emitted:', len(emitted_pages), sorted(emitted_pages))
print('Cards with verification flattened:', list(attest.keys()))
if multi_warn:
    print('WARNING multi-attest cards (only first kept):', sorted(set(multi_warn)))
