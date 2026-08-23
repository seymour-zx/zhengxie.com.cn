# -*- coding: utf-8 -*-
"""向各频道 self_links.xlsx 追加「参考」分类行（用于佐证收录/分类正确性）。
每个参考条目：分类='参考'，links 列承载 URL。站序从最大值顺延（不参与正常导航但单列分组）。
"""
import openpyxl
from openpyxl import Workbook

BASE = r"D:\Universal Space\zhengxie.com.cn\directory"

# 各频道参考来源（如实标注：依据公开权威信息整理）
REFS = {
    "nav": [
        ("中国互联网络信息中心(CNNIC)", "https://www.cnnic.net.cn/", "国内互联网基础资源与网站收录权威机构"),
        ("工业和信息化部 ICP/IP 备案系统", "https://beian.miit.gov.cn/", "核验站点备案与主办单位性质的权威入口"),
    ],
    "gov": [
        ("中国政府网·国务院组织机构", "https://www.gov.cn/", "国务院组成部门与机构序列的权威发布源"),
        ("中央机构编制委员会办公室", "http://www.scopsr.gov.cn/", "机关事业单位机构编制权威核定"),
    ],
    "npc": [
        ("中国政府网·国家机构", "https://www.gov.cn/", "国家权力机关序列的权威索引，佐证人大机构归类"),
    ],
    "party": [
        ("共产党员网", "https://www.12371.cn/", "中共中央组织部主管的党建权威门户"),
        ("中国共产党新闻网", "http://cpc.people.com.cn/", "党的理论创新与组织建设权威发布源"),
    ],
    "zhengxie": [
        ("全国政协网", "http://www.cppcc.gov.cn/", "中国人民政治协商会议全国委员会官方门户"),
        ("中央社会主义学院", "http://www.cnss.org.cn/", "民主党派与统一战线人才培养主阵地"),
    ],
    "world-gov": [
        ("CIA World Factbook", "https://www.cia.gov/the-world-factbook/", "各国政府体制与官方门户的国际参考索引"),
        ("联合国", "https://www.un.org/", "主权国家与政府承认的权威参考"),
    ],
}

def append_refs():
    for ch, refs in REFS.items():
        path = f"{BASE}/{ch}/assets/xlsx/self_links.xlsx"
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # 找表头与现有最大站序
        header = None
        max_seq = 0
        rows = []
        for i, r in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 1:
                header = list(r)
                continue
            rows.append(list(r))
            try:
                seq = int(r[0]) if r[0] is not None else 0
                if seq > max_seq:
                    max_seq = seq
            except (ValueError, TypeError):
                pass
        # 重组：过滤掉已有的「参考」分类（避免重复追加）
        kept = [r for r in rows if str(r[1]).strip() != "参考"]
        next_seq = max_seq
        for name, url, note in refs:
            next_seq += 1
            kept.append([next_seq, "参考", 3, name, note, "", "", f"来源,{url}"])
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = ws.title
        ws2.append(header)
        for r in kept:
            ws2.append(r)
        wb2.save(path)
        print(f"{ch}: 追加 {len(refs)} 条参考，当前最大站序 {next_seq}")

append_refs()
